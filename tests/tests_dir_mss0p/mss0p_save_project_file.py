import numpy as np
import openpyxl as pxl
import os, sys
from datetime import datetime

# Given four dictionaries containing information about a mss0p project:
# basic_info, camera_parameters, pois_definition, and image_sources,
# a path to a mss0p project file (project_file_path),
# a text widget (print_widget) for displaying messages,
# if project_file_path is None or an empty string, this function will 
# use tkinter file dialog to ask the user to select an xlsx file path to write. 
# For the basic_info dictionary, this function saves the information to an xlsx file.
# basic_info contains keys like 'num_cameras', 'num_pois', and 'num_steps', and 
# the value should be an integer, a float, or a string, which can simply be 
# converted to a string and written to a cell in the xlsx file.
# If basic_info is None, it will create an empty sheet named 'basic_info' in the xlsx file.
# For example, if basic_info is:
#   basic_info = {'num_cameras': 4, 'num_pois': 10, 'num_steps': 100}
# The xlsx file will have a sheet named 'basic_info' with the following content:
#  | key              | value          |
#  |------------------|----------------|
#  | num_cameras      | 4              |
#  | num_pois         | 10             |
#  | num_steps        | 100            |
#  |                  |                |
#  / ------------ \
#  | basic_info   |
#
# Worksheet "camera_parameters"
#  The worksheet named "camera_parameters" contains camera parameters for each camera.
#  The first row is the name of the cameras (except the first column which is a constant "parameter_name").
#  For example, the first row (starting from the second column) could be:
#  camera_name_1, camera_name_2, camera_name_3, camera_name_4
#  You can change camera_name_1, ..., to any names you want. Just avoid using double underline (__).
#  The first column is the parameter name. From the 2nd row, they are image_width, image_height, 
#  rvec_x, rvec_y, rvec_z, 
#  tvec_x, tvec_y, tvec_z, cmat11_fx, cmat12, cmat13_cx, cmat21, cmat22_fy, cmat23_cy, cmat31, cmat32, cmat33,
#  k1, k2, k3, p1, p2, s1, s2, s3, s4, tau_x, tau_y 
#  The worksheet "camera_parameters" should look like this:
# 
# | parameter_name | camera_name_1  | camera_name_2  | camera_name_3  | camera_name_4  |
# |----------------|----------------|----------------|----------------|----------------|
# | image_width    | 1920           | 1920           | 1920           | 1920           |
# | image_height   | 1080           | 1080           | 1080           | 1080           |
# | rvec_x         | 0.0            | 0.0            | 0.0            | 0.0            |
# | rvec_y         | 0.0            | 0.0            | 0.0            | 0.0            |
# | rvec_z         | 0.0            | 0.0            | 0.0            | 0.0            |
# | tvec_x         | 0.0            | 0.0            | 0.0            | 0.0            |
# | tvec_y         | 0.0            | 0.0            | 0.0            | 0.0            |
# | tvec_z         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat11_fx      | 1000.0         | 1000.0         | 1000.0         | 1000.0         |
# | cmat12         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat13_cx      | 960.0          | 960.0          | 960.0          | 960.0          |
# | cmat21         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat22_fy      | 1000.0         | 1000.0         | 1000.0         | 1000.0         |
# | cmat23_cy      | 540.0          | 540.0          | 540.0          | 540.0          |
# | cmat31         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat32         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat33         | 1.0            | 1.0            | 1.0            | 1.0            |
# | k1             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k2             | 0.0            | 0.0            | 0.0            | 0.0            |
# | p1             | 0.0            | 0.0            | 0.0            | 0.0            |
# | p2             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k3             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k4             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k5             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k6             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s1             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s2             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s3             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s4             | 0.0            | 0.0            | 0.0            | 0.0            |
# | tau_x          | 0.0            | 0.0            | 0.0            | 0.0            |
# | tau_y          | 0.0            | 0.0            | 0.0            | 0.0            |
#
#  / ------------------- \
#  | camera_parameters   | 
# The camera_paramters dictionary looks like this:
# camera_parameters = {
#     'camera_name_1': {
#         'image_size': (1920, 1080),
#         'rvec': np.array([0.0, 0.0, 0.0]).reshape(3,1),  # rotation vector as a 3x1 array
#         'tvec': np.array([0.0, 0.0, 0.0]).reshape(3,1),  # translation vector as a 3x1 array
#         'cmat': np.array([[1000.0, 0.0, 960.0], [0.0, 1000.0, 540.0], [0.0, 0.0, 1.0]]).reshape(3, 3),
#         'dvec': np.array([0.0, 0.0, 0.0, 0.0]).reshape(-1, 1)  # or more elements if available
#     }
#    'camera_name_2': {
#         'image_size': (1920, 1080),
#         'rvec': np.array([0.0, 0.0, 0.0]).reshape(3,1),  # rotation vector as a 3-element array
#         'tvec': np.array([0.0, 0.0, 0.0]).reshape(3,1),  # translation vector as a 3-element array
#         'cmat': np.array([[1000.0, 0.0, 960.0], [0.0, 1000.0, 540.0], [0.0, 0.0, 1.0]]).reshape(3, 3),
#         'dvec': np.array([0.0, 0.0, 0.0, 0.0]).reshape(-1, 1)  # or more elements if available
# #   }
# } 
# If dvec length is less than 14, the remaining elements are written as 0.0.
#
# Worksheet "pois_definition"
# The worksheet named "pois_definition" contains the definition of points of interest (POIs).
# Each poi has the following properties:
# - xw, yw, zw: the world coordinates of the poi, which is a 3-element array 
# - xi, yi: the image coordinates of the poi in each camera, which is a 2-element array
# - x0, y0, w, h: the template of the poi in each camera, which is a 4-element integer array
# 
#  The worksheet "pois_definition" should look like this. 
#  You can replace poi_name_1, ..., etc., with any name you want, just avoid using double underline (__).
#  You can replace camera_name_1, ..., etc., with any name you want, just avoid using double underline (__), 
#  and the camera_name_1, ..., etc., should match the camera names in the camera_parameters worksheet.
#  In the header row, starting from the 5th column, the property names are prefixed with "xi__", "yi__", "x0__", "y0__", "w__", "h__"
#  followed by the camera name, for example, "xi__camera_name_1", "yi__camera_name_1", "x0__camera_name_1", "y0__camera_name_1", "w__camera_name_1", "h__camera_name_1"
#  The separator between the property name and the camera name is a double underline (__).
#  That is why the camera names in the camera_parameters worksheet should not contain double underline (__).
# 
# | poi_name         | xw   | yw   | zw   | xi__camera_name_1  | yi__camera_name_1  | x0__camera_name_1  | y0__camera_name_1  | w__camera_name_1  | h__camera_name_1  | xi_ _camera_name_2 | yi__camera_name_2  |
# |------------------|------|------|------|--------------------|--------------------|--------------------|--------------------|-------------------|-------------------|--------------------|--------------------|
# | poi_name_1       | 1.0  | 2.0  | 3.0  | 100.0              | 200.0              | 50                 | 50                 | 100               | 100               | 150.0              | 250.0              |
# | poi_name_2       | 4.0  | 5.0  | 6.0  | 110.0              | 210.0              | 60                 | 60                 | 120               | 120               | 160.0              | 260.0              |
# | poi_name_3       | 7.0  | 8.0  | 9.0  | 120.0              | 220.0              | 70                 | 70                 | 140               | 140               | 170.0              | 270.0              |
#
#  / ------------------- \
#  | pois_definition     | 
#
#  The pois_definition dictionary looks like this:
#  pois_definition = {
#      'poi_name_1': {
#          'Xw': np.array([1.0, 2.0, 3.0]),  # world coordinates of the poi
#          'Xi': {
#              'camera_name_1': np.array([100.0, 200.0]),  # image coordinates of the poi in camera 1
#              'camera_name_2': np.array([150.0, 250.0]),  # image coordinates of the poi in camera 2
#          },
#          'Tmplt': {
#              'camera_name_1': np.array([50, 50, 100, 100], dtype=np.int32),  # template in camera 1
#              'camera_name_2': np.array([60, 60, 120, 120], dtype=np.int32),  # template in camera 2
#          }
#      }
#      'poi_name_2': {
#          'Xw': np.array([4.0, 5.0, 6.0]),  # world coordinates of the poi
#          'Xi': {
#              'camera_name_1': np.array([110.0, 210.0]),  # image coordinates of the poi in camera 1
#              'camera_name_2': np.array([160.0, 260.0]),  # image coordinates of the poi in camera 2
#          },
#          'Tmplt': {
#              'camera_name_1': np.array([60, 60, 120, 120], dtype=np.int32),  # template in camera 1
#              'camera_name_2': np.array([70, 70, 140, 140], dtype=np.int32),  # template in camera 2
#          }
#      }
#  }
#  The pois_definition dictionary can have any number of POIs, and each POI can have any number of cameras.
# 
#  # Worksheet "image_sources"
#  The image_sources dictionary looks like this:
#  image_sources = {
#      'camera_name_1': '/path/to/video1.mp4',
#      'camera_name_2': '/path/to/video2.mp4',
#      'camera_name_3': ['/path/to/camera_name_3/image1.jpg', '/path/to/camera_name_3/image2.jpg', ...]
#      'camera_name_4': ['/path/to/camera_name_4/image1.jpg', '/path/to/camera_name_4/image2.jpg', ...]
#  }

##  The worksheet "image_sources" should look like this:
# | camera_name      | image_source_file_path  |
# |------------------|-------------------------|
# | camera_name_1    | relative_path/to/video1.mp4     |
# | camera_name_2    | relative_path/to/video2.mp4     |
# | camera_name_3    | relative_path/to/camera_name_3/file_list_camera3.txt |
# | camera_name_4    | relative_path/to/camera_name_4/file_list_camera4.txt |
#  / ------------------- \
#  | image_sources     | 
#  
#  The relative path is relative to the project file path (project_file_path).
#  For example, if the project file is saved in /path/to/project/project.xlsx,
#  and the image source file path is /path/to/project/camera_name_1/video1.mp4,
#  then the image source file path in the xlsx file will be ../camera_name_1/video1.mp4.

def save_project_file(basic_info=None, camera_parameters=None, pois_definition=None, 
                            image_sources=None, project_file_path=None, print_widget=None):
    # if project_file_path is None or an empty string, this function will 
    # # use tkinter file dialog to ask the user to select an xlsx file path to write.
    if project_file_path is None or project_file_path == "":
        from tkinter import Tk
        from tkinter.filedialog import asksaveasfilename
        Tk().withdraw()
        project_file_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Project File"
        )
    # if user cancels the dialog, returns without saving
    if not project_file_path:
        if print_widget:
            print_widget.insert("end", "No file selected. Project not saved.\n")
        return
    # Create a new workbook and add sheets
    workbook = pxl.Workbook()
    # Add basic_info sheet
    if basic_info is None:
        basic_info = {}
    basic_info_sheet = workbook.active
    basic_info_sheet.title = "basic_info"
    # Write basic_info to the sheet
    if basic_info:
        basic_info_sheet.append(["key", "value"])
        for key, value in basic_info.items():
            basic_info_sheet.append([key, str(value)])
#    else:
#        basic_info_sheet.append(["No basic info provided."])
    #
    # Add camera_parameters sheet
    camera_parameters_sheet = workbook.create_sheet(title="camera_parameters")
    if camera_parameters is None:
        camera_parameters = {}
    if camera_parameters:
        # Get the camera names from the camera_parameters dictionary
        camera_names = list(camera_parameters.keys())
        if camera_names:
            # Write the header row with camera names
            header = ["parameter_name"] + camera_names
            camera_parameters_sheet.append(header)
            # Write each parameter for each camera
            parameters = [
                "image_width", "image_height", "rvec_x", "rvec_y", "rvec_z",
                "tvec_x", "tvec_y", "tvec_z", "cmat11_fx", "cmat12", "cmat13_cx",
                "cmat21", "cmat22_fy", "cmat23_cy", "cmat31", "cmat32", "cmat33",
                "k1", "k2", "p1", "p2", "k3", "k4", "k5", "k6", "s1", "s2", "s3", "s4", 
                "tau_x", "tau_y"
            ]
            for param in parameters:
                row = [param]
                for camera_name in camera_names:
                    if camera_name in camera_parameters:
                        cam_params = camera_parameters[camera_name]
                        if param == 'image_width':
                            row.append(str(cam_params['image_size'][0]))
                        elif param == 'image_height':
                            row.append(str(cam_params['image_size'][1]))
                        elif param.startswith('rvec_x'):
                            row.append(str(cam_params['rvec'][0, 0]))
                        elif param.startswith('rvec_y'):
                            row.append(str(cam_params['rvec'][1, 0]))
                        elif param.startswith('rvec_z'):
                            row.append(str(cam_params['rvec'][2, 0]))
                        elif param.startswith('tvec_x'):
                            row.append(str(cam_params['tvec'][0, 0]))
                        elif param.startswith('tvec_y'):
                            row.append(str(cam_params['tvec'][1, 0]))
                        elif param.startswith('tvec_z'):
                            row.append(str(cam_params['tvec'][2, 0]))
                        elif param.startswith('cmat11_fx'):
                            row.append(str(cam_params['cmat'][0, 0]))
                        elif param.startswith('cmat12'):
                            row.append(str(cam_params['cmat'][0, 1]))
                        elif param.startswith('cmat13_cx'):
                            row.append(str(cam_params['cmat'][0, 2]))
                        elif param.startswith('cmat21'):
                            row.append(str(cam_params['cmat'][1, 0]))
                        elif param.startswith('cmat22_fy'):
                            row.append(str(cam_params['cmat'][1, 1]))
                        elif param.startswith('cmat23_cy'):
                            row.append(str(cam_params['cmat'][1, 2]))
                        elif param.startswith('cmat31'):
                            row.append(str(cam_params['cmat'][2, 0]))
                        elif param.startswith('cmat32'):
                            row.append(str(cam_params['cmat'][2, 1]))
                        elif param.startswith('cmat33'):
                            row.append(str(cam_params['cmat'][2, 2]))
                        elif param.startswith('k1'):
                            row.append(str(cam_params['dvec'][0, 0])) if len(cam_params['dvec']) > 0 else row.append('0') 
                        elif param.startswith('k2'):
                            row.append(str(cam_params['dvec'][1, 0])) if len(cam_params['dvec']) > 1 else row.append('0') 
                        elif param.startswith('p1'):
                            row.append(str(cam_params['dvec'][2, 0])) if len(cam_params['dvec']) > 2 else row.append('0') 
                        elif param.startswith('p2'):
                            row.append(str(cam_params['dvec'][3, 0])) if len(cam_params['dvec']) > 3 else row.append('0') 
                        elif param.startswith('k3'):
                            row.append(str(cam_params['dvec'][4, 0])) if len(cam_params['dvec']) > 4 else row.append('0') 
                        elif param.startswith('k4'):
                            row.append(str(cam_params['dvec'][5, 0])) if len(cam_params['dvec']) > 5 else row.append('0') 
                        elif param.startswith('k5'):
                            row.append(str(cam_params['dvec'][6, 0])) if len(cam_params['dvec']) > 6 else row.append('0') 
                        elif param.startswith('k6'):
                            row.append(str(cam_params['dvec'][7, 0])) if len(cam_params['dvec']) > 7 else row.append('0') 
                        elif param.startswith('s1'):
                            row.append(str(cam_params['dvec'][8, 0])) if len(cam_params['dvec']) > 8 else row.append('0') 
                        elif param.startswith('s2'):
                            row.append(str(cam_params['dvec'][9, 0])) if len(cam_params['dvec']) > 9 else row.append('0') 
                        elif param.startswith('s3'):
                            row.append(str(cam_params['dvec'][10, 0])) if len(cam_params['dvec']) > 10 else row.append('0')
                        elif param.startswith('s4'):
                            row.append(str(cam_params['dvec'][11, 0])) if len(cam_params['dvec']) > 11 else row.append('0')
                        elif param.startswith('tau_x'):
                            row.append(str(cam_params['dvec'][12, 0])) if len(cam_params['dvec']) > 12 else row.append('0')
                        elif param.startswith('tau_y'):
                            row.append(str(cam_params['dvec'][13, 0])) if len(cam_params['dvec']) > 13 else row.append('0')
                        else:
                            row.append('0')
                        # end of if param is 'image_width', 'image_height', etc.
                    # end of if camera_name in camera_parameters (of course it will be)
                # end of for camera_name in camera_names
                camera_parameters_sheet.append(row)
            # end of for param in parameters
        # end of if camera_names
    # end of if camera_parameters is not None
# 
    # Add pois_definition sheet
    pois_definition_sheet = workbook.create_sheet(title="pois_definition")
    if pois_definition is None:
        pois_definition = {}
    if pois_definition:
        # Write the header row
        header = ["poi_name", "xw", "yw", "zw"]
        camera_names = list(camera_parameters.keys()) if camera_parameters else []
        for camera_name in camera_names:
            header.extend([
                f"xi__{camera_name}", f"yi__{camera_name}",
                f"x0__{camera_name}", f"y0__{camera_name}",
                f"w__{camera_name}", f"h__{camera_name}"
            ])
        pois_definition_sheet.append(header)
        # Write each poi's data
        for poi_name, poi_data in pois_definition.items():
            row = [poi_name, poi_data['Xw'][0], poi_data['Xw'][1], poi_data['Xw'][2]]
            for camera_name in camera_names:
                if camera_name in poi_data['Xi']:
                    xi = poi_data['Xi'][camera_name]
                    x0 = poi_data['Tmplt'][camera_name]
                    row.extend([xi[0], xi[1], x0[0], x0[1], x0[2], x0[3]])
                else:
                    row.extend(['', '', '', '', '', ''])
            # end of for camera_name in camera_names
            pois_definition_sheet.append(row)
        # end of for poi_name, poi_data in pois_definition.items()
    # end of if pois_definition is not None
#
    # Add image_sources sheet
    image_sources_sheet = workbook.create_sheet(title="image_sources")
    if image_sources is None:
        image_sources = {}
    if image_sources:
        # Write the header row
        image_sources_sheet.append(["camera_name", "image_source_file_path"])
        # Write each camera's image source
        for camera_name, image_source in image_sources.items():
            if isinstance(image_source, list):
                # If the image source is a list, save it as a text file and write the path
                file_list_path = os.path.join(os.path.dirname(project_file_path), f"file_list_{camera_name}.txt")
                with open(file_list_path, 'w') as f:
                    for img in image_source:
                        f.write(f"{img}\n")
                image_sources_sheet.append([camera_name, file_list_path])
            else:
                # If it's a single file path, write it directly
                relative_path = os.path.relpath(image_source, os.path.dirname(project_file_path))
                image_sources_sheet.append([camera_name, relative_path])
            # end of if isinstance(image_source, list)
        # end of for camera_name, image_source in image_sources.items()
    # end of if image_sources is not None

    # save the workbook to the specified file path
    workbook.save(project_file_path)
    if print_widget:
        print_widget.insert("end", f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
        print_widget.insert("end", f"Project file saved to: {project_file_path}\n")
    else:
        print(f"Project file saved to: {project_file_path}")

    # end of function save_mss0p_project_file
    # Return the path of the saved project file
    return project_file_path
