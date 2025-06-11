import numpy as np
import openpyxl as pxl
import os, sys
from datetime import datetime

# Given a path to a mss0p project file (project_file_path), 
# which is an xlsx file with a specific structure, 
# this function loads the project file and returns four dictionaries:
#   basic_info, 
#   camera_parameters, 
#   pois_definition, and 
#   image_sources.
# a mss0p project file is a .xlsx file with a specific structure
# Worksheet "basic_info" 
#  The worksheet named "basic_info" contains basic information about the project.
#  In basic_info, the first column is the key and the second column is the value.
#  For example: num_cameras , 4 means the project has 4 cameras.
#  For example: num_pois , 10 means the project has 10 points of interest.
#  For example: num_steps , 100 means the project has 100 steps.
#  The worksheet "basic_info" should look like this:
#
#  | key              | value          |
#  |------------------|----------------|
#  | num_cameras      | 4              |
#  | num_pois         | 10             |
#  | num_steps        | 100            |
#  |                  |                |
#  |------------------|----------------|
#  / ------------ \
#  | basic_info   |
#  ------------------------------------ 
#  It transforms to a dictionary like this:
#  basic_info = {'num_cameras': 4, 'num_pois': 10, 'num_steps': 100}
#  For example:
#  num_cameras = basic_info['num_cameras']  # number of cameras
#
#
# Worksheet "camera_parameters"
#  The worksheet named "camera_parameters" contains camera parameters for each camera.
#  The first row is the name of the cameras (except the first column which is a constant "parameter_name").
#  For example, the first row (starting from the second column) could be:
#  camera_name_1, camera_name_2, camera_name_3, camera_name_4
#  You can change camera_name_1, ..., to any names you want. Just avoid using double underline (__).
#  The first column is the parameter name. From the 2nd row, they are image_width, image_height, 
#  rvec_x, rvec_y, rvec_z, 
#  tvec_x, tvec_y, tvec_z, cmat11_fx, cmat12, cmat13_cx, cmat21, cmat22_fy, cmat23_cy, cmat31, cmat32, cmat33,
#  k1, k2, k3, p1, p2, s1, s2, s3, s4, tau_x, tau_y 
#  The worksheet "camera_parameters" should look like this:
# 
# | parameter_name | camera_name_1  | camera_name_2  | camera_name_3  | camera_name_4  |
# |----------------|----------------|----------------|----------------|----------------|
# | image_width    | 1920           | 1920           | 1920           | 1920           |
# | image_height   | 1080           | 1080           | 1080           | 1080           |
# | rvec_x         | 0.0            | 0.0            | 0.0            | 0.0            |
# | rvec_y         | 0.0            | 0.0            | 0.0            | 0.0            |
# | rvec_z         | 0.0            | 0.0            | 0.0            | 0.0            |
# | tvec_x         | 0.0            | 0.0            | 0.0            | 0.0            |
# | tvec_y         | 0.0            | 0.0            | 0.0            | 0.0            |
# | tvec_z         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat11_fx      | 1000.0         | 1000.0         | 1000.0         | 1000.0         |
# | cmat12         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat13_cx      | 960.0          | 960.0          | 960.0          | 960.0          |
# | cmat21         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat22_fy      | 1000.0         | 1000.0         | 1000.0         | 1000.0         |
# | cmat23_cy      | 540.0          | 540.0          | 540.0          | 540.0          |
# | cmat31         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat32         | 0.0            | 0.0            | 0.0            | 0.0            |
# | cmat33         | 1.0            | 1.0            | 1.0            | 1.0            |
# | k1             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k2             | 0.0            | 0.0            | 0.0            | 0.0            |
# | p1             | 0.0            | 0.0            | 0.0            | 0.0            |
# | p2             | 0.0            | 0.0            | 0.0            | 0.0            |
#  (below are optional)
# | k3             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k4             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k5             | 0.0            | 0.0            | 0.0            | 0.0            |
# | k6             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s1             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s2             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s3             | 0.0            | 0.0            | 0.0            | 0.0            |
# | s4             | 0.0            | 0.0            | 0.0            | 0.0            |
# | tau_x          | 0.0            | 0.0            | 0.0            | 0.0            |
# | tau_y          | 0.0            | 0.0            | 0.0            | 0.0            |
#
#  / ------------------- \
#  | camera_parameters   | 
# It transforms to a dictionary like this:
# camera_parameters = {
#     'camera_name_1': {
#         'image_size': (1920, 1080),
#         'rvec': np.array([0.0, 0.0, 0.0]).reshape(3,1),  # rotation vector as a 3-element array
#         'tvec': np.array([0.0, 0.0, 0.0]).reshape(3,1),  # translation vector as a 3-element array
#         'cmat': np.array([[1000.0, 0.0, 960.0], [0.0, 1000.0, 540.0], [0.0, 0.0, 1.0]]).reshape(3, 3),,
#         'dvec': np.array([0.0, 0.0, 0.0, 0.0]).reshape(-1, 1)  # or more elements if available
#     }
# For example:
#   rvec = camera_parameters['camera_name_1']['rvec']  # a 3x1 numpy array
#
# Worksheet "pois_definition"
# The worksheet named "pois_definition" contains the definition of points of interest (POIs).
# Each poi has the following properties:
# - xw, yw, zw: the world coordinates of the poi, which is a 3-element array 
# - xi, yi: the image coordinates of the poi in each camera, which is a 2-element array
# - x0, y0, w, h: the template of the poi in each camera, which is a 4-element integer array
# 
#  The worksheet "pois_definition" should look like this. 
#  You can replace poi_name_1, ..., etc., with any name you want, just avoid using double underline (__).
#  You can replace camera_name_1, ..., etc., with any name you want, just avoid using double underline (__), 
#  and the camera_name_1, ..., etc., should match the camera names in the camera_parameters worksheet.
#  In the header row, starting from the 5th column, the property names are prefixed with "xi__", "yi__", "x0__", "y0__", "w__", "h__"
#  followed by the camera name, for example, "xi__camera_name_1", "yi__camera_name_1", "x0__camera_name_1", "y0__camera_name_1", "w__camera_name_1", "h__camera_name_1"
#  The separator between the property name and the camera name is a double underline (__).
#  That is why the camera names in the camera_parameters worksheet should not contain double underline (__).
# 
# | poi_name         | xw   | yw   | zw   | xi__camera_name_1  | yi__camera_name_1  | x0__camera_name_1  | y0__camera_name_1  | w__camera_name_1  | h__camera_name_1  | xi_ _camera_name_2 | yi__camera_name_2  |
# |------------------|------|------|------|--------------------|--------------------|--------------------|--------------------|-------------------|-------------------|--------------------|--------------------|
# | poi_name_1       | 1.0  | 2.0  | 3.0  | 100.0              | 200.0              | 50                 | 50                 | 100               | 100               | 150.0              | 250.0              |
# | poi_name_2       | 4.0  | 5.0  | 6.0  | 110.0              | 210.0              | 60                 | 60                 | 120               | 120               | 160.0              | 260.0              |
# | poi_name_3       | 7.0  | 8.0  | 9.0  | 120.0              | 220.0              | 70                 | 70                 | 140               | 140               | 170.0              | 270.0              |
#
#  / ------------------- \
#  | pois_definition     | 
#
#  It transforms to a dictionary like this:
# The data format of pois_definition dictionary is:
#  pois_definition = {
#      'poi_name_1': {
#          'Xw': (1.0, 2.0, 3.0),  # world coordinates of the poi
#          'Xi': {
#              'camera_name_1': (100.0, 200.0),  # image coordinates of the poi in camera 1
#              'camera_name_2': (150.0, 250.0),  # image coordinates of the poi in camera 2
#          },
#          'Tmplt': {
#              'camera_name_1': (50, 50, 100, 100),  # template in camera 1
#              'camera_name_2': (60, 60, 120, 120),  # template in camera 2
#          }
#      }
#      'poi_name_2': {
#          'Xw': (4.0, 5.0, 6.0),  # world coordinates of the poi
#          'Xi': {
#              'camera_name_1': (110.0, 210.0),  # image coordinates of the poi in camera 1
#              'camera_name_2': (160.0, 260.0),  # image coordinates of the poi in camera 2
#          },
#          'Tmplt': {
#              'camera_name_1': (60, 60, 120, 120),  # template in camera 1
#              'camera_name_2': (70, 70, 140, 140),  # template in camera 2
#          }
#      }
#  }
#  where 'poi_name_1' and 'poi_name_2' are the names of the points of interest (POIs), 
#  and can be changed to any string.
#  where 'camera_name_1' and 'camera_name_2' are the names of the cameras,
#  and can be changed to any string.
#  where 'Xw' is the world coordinates of the POI,
#  'Xi' is a dictionary of image coordinates of the POI in different cameras,
#  and 'Tmplt' is a dictionary of templates (a small region of interest, ROI of the template)
#  in different cameras.
#  The words 'Xw', 'Xi', and 'Tmplt' are fixed and should not be changed.
#  For example:
#    xw = pois_definition['poi_name_1']['Xw']  # a 3-element tuple
#    xi = pois_definition['poi_name_1']['Xi']['camera_name_1']  # a 2-element tuple
#    tmplt = pois_definition['poi_name_1']['Tmplt']['camera_name_1']  # a 4-element integer tuple (x0, y0, w, h)
# 
#
# Worksheet "image_sources"
# The worksheet named "image_sources" contains the image sources for each camera.
#  Each camera has a video file or a text file that a list of image files, which are the images taken by the camera.
#  If the file has an extension .txt, it is a text file that contains a list of image files.
#    The file-list text file looks like this:
#       /path/to/image1.jpg
#       /path/to/image2.jpg
#       /path/to/image3.jpg
#  If the file has an extension .mp4, .avi, .mov, etc., it is a video file that contains the images taken by the camera.
#  If the file path is a relative path, it is relative to the path of the file-list text file.  
#  If the file path is an absolute path, it is the absolute path to the video file or the text file.
#  
#  The worksheet "image_sources" should look like this:
# | camera_name      | image_source_file_path  |
# |------------------|-------------------------|
# | camera_name_1    | /path/to/video1.mp4     |
# | camera_name_2    | /path/to/video2.mp4     |
# | camera_name_3    | /abs_path/to/camera_name_3/file_list_camera3.txt |
# | camera_name_4    | relative_path/to/camera_name_4/file_list_camera4.txt |
#  / ------------------- \
#  | image_sources     | 
# 
#  If the image source file path is a relative path, it is relative to the project file path 
#  (variable name is project_file_path), 
#  If the image source file path is an absolute path, it is the absolute path to the video file or the text file.
#   
#  It transforms to a dictionary like this:
#  image_sources = {
#      'camera_name_1': '/path/to/video1.mp4',
#      'camera_name_2': '/path/to/video2.mp4',
#      'camera_name_3': ['/path/to/camera_name_3/image1.jpg', '/path/to/camera_name_3/image2.jpg', ...]
#      'camera_name_4': ['/path/to/camera_name_4/image1.jpg', '/path/to/camera_name_4/image2.jpg', ...]
#  }
#  If the file path in the  file path is a relative path, it is relative to the project file path (project_file_path).
#  If the image source file path is an absolute path, it is the absolute path to the video file or the text file.

def load_project_file(project_file_path=None, print_widget=None):
    # Four return values are initialized to None
    basic_info, pois_definition, camera_parameters, image_sources = None, None, None, None
    # open the project file 
    if project_file_path is None:
        # if project_file_path is not provided, pops up a tk file dialog to ask user to select the project file (an xlsx file)
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()  # Hide the root window
        filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
        project_file_path = filedialog.askopenfilename(
            title="Select a project file",
            filetypes=filetypes
        )
        # if user clicks cancel, this function returns None
        if not project_file_path:
            # print_widget is a Text widget, if provided, print a message to it
            if print_widget is not None:
                # print date/time and a warning message to the print_widget
                print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                print_widget.insert(tk.END, "#   Warning: No project file selected. Returning None for all variables.\n")
            return None, None, None, None
    # check if the project file exists
    if not os.path.isfile(project_file_path):
        # if file does not exists, 
        # print date/time and a warning message to the print_widget
        if print_widget is not None:
            print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
            print_widget.insert(tk.END, f"#   Warning: The project file {project_file_path} does not exist. Returning None for all variables.\n")
        return None, None, None, None
    # check if the project file is an xlsx file 
    if not project_file_path.endswith('.xlsx'):
        if print_widget is not None:
            print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
            print_widget.insert(tk.END, f"# Error: The project file {project_file_path} is not a valid .xlsx file.")
        return None, None, None, None
    # open the project file which is an xlsx file
    try:
        wb = pxl.load_workbook(project_file_path, data_only=True)
    except:
        if print_widget is not None:
            print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
            print_widget.insert(tk.END, f"# Error: Failed to open the project file {project_file_path}. It may not be a valid .xlsx file.\n")
        return None, None, None, None
        
    # get the basic_info worksheet
    if 'basic_info' not in wb.sheetnames:
        # print warning message in the print_widget if provided
        if print_widget is not None: 
            print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
            print_widget.insert(tk.END, "# Warning: The project file does not contain a 'basic_info' worksheet.\n")
    else:
        ws_basic_info = wb['basic_info']
        # read the basic information from the basic_info worksheet
        basic_info = {}
        for row in ws_basic_info.iter_rows(min_row=2, max_col=2, values_only=True):
            key, value = row
            basic_info[key.strip()] = value.strip() if isinstance(value, str) else value
            # print date/time, and the key and value to the print_widget if provided
            if print_widget is not None:
                print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                print_widget.insert(tk.END, f"#   {key.strip()}: {value.strip() if isinstance(value, str) else value}\n")
    # camera_parameters: 
    # get the camera_parameters worksheet
    # for example, the headers in the camera_parameters worksheet could be:
    # parameter_name, camera_name_1, camera_name_2, camera_name_3, camera_name_4
    # the 2nd row could be:
    # image_width, 1920, 1920, 1920, 1920
    # the 3rd row could be:
    # image_height, 1080, 1080, 1080, 1080
    # then, camera_parameters['camera_name_1'] will be a dictionary with keys 'image_size', 'rvec', 'tvec', 'cmat', 'dvec'
    # That is:
    # camera_parameters['camera_name_1']['rvec'] will be a 3-element array with the rotation vector
    # camera_parameters['camera_name_1']['tvec'] will be a 3-element array with the translation vector
    # camera_parameters['camera_name_1']['cmat'] will be a 3x3 array with the camera matrix
    # camera_parameters['camera_name_1']['dvec'] will be a 12-element array with the distortion vector
    # the ['dvec'] can have different lengths (i.e., 4, 5, 8, 12, or 14), depending on the camera model used. 
    if 'camera_parameters' not in wb.sheetnames:
        print("# Error: The project file does not contain a 'camera_parameters' worksheet.")
        # if print_widget is provided, print date/time, and a message to it
        if print_widget is not None:
            print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
            print_widget.insert(tk.END, "#   Error: The project file does not contain a 'camera_parameters' worksheet.\n")
    else:
        ws_camera_parameters = wb['camera_parameters']
        # read the camera parameters from the camera_parameters worksheet
        camera_parameters = {}
        headers = [cell.value for cell in ws_camera_parameters[1]]
        cam_names = headers[1:]
        # strip all strings in cam_names
        cam_names = [name.strip() for name in cam_names if isinstance(name, str)]
        # if cam_names contains double underline (__), print a warning message, and replace it with a single underline (_) 
        for i, name in enumerate(cam_names):
            if '__' in name:
                print(f"# Warning: Camera name '{name}' contains double underline. Replacing it with a single underline.")
                cam_names[i] = name.replace('__', '_')
                # if print_widget is provided, print date/time, and a message to it
                if print_widget is not None:
                    print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                    print_widget.insert(tk.END, f"#   Warning: Camera name '{name}' contains double underline. Replacing it with a single underline.\n")
        # print date/time, and the camera names to the print_widget if provided
        if print_widget is not None:
            print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
            print_widget.insert(tk.END, "#   Number of cameras: " + str(len(cam_names)) + "\n")
            print_widget.insert(tk.END, "#   Camera names: " + ', '.join(cam_names) + "\n")                
        # create a numpy array to hold the camera parameters
        num_cameras = len(cam_names)
        cam_parameters_nparray = np.zeros((31, num_cameras), dtype=np.float64)  # 31 parameters
        # read the camera parameters from the camera_parameters worksheet
        i = 0
        for row in ws_camera_parameters.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            parameter_name = row[0]
            values = row[1:]
            cam_parameters_nparray[i, :] = np.array(values, dtype=np.float64)
            i += 1
        # convert the camera parameters numpy array to a dictionary
        # for example, camera_parameters['camera_name_1'] will be a dictionary with keys 'image_size', 'rvec', 'tvec', 'cmat', 'dvec'
        # where 'image_size' is a tuple (width, height), 'rvec' is a 3-element array, 'tvec' is a 3-element array,
        for icam in range(num_cameras):
            camera_parameters[cam_names[icam]] = {\
                'image_size': (int(cam_parameters_nparray[0, icam]), int(cam_parameters_nparray[1, icam])),
                'rvec': cam_parameters_nparray[2:5, icam].reshape(3,1),  # rvec_x, rvec_y, rvec_z,
                'tvec': cam_parameters_nparray[5:8, icam].reshape(3,1),  # tvec_x, tvec_y, tvec_z
                'cmat': cam_parameters_nparray[8:17, icam].reshape((3, 3)),  # cmat11_fx, cmat12, cmat13_cx, cmat21, cmat22_fy, cmat23_cy, cmat31, cmat32, cmat33
                'dvec': cam_parameters_nparray[17:, icam].reshape(-1,1),  # k1, k2, k3, p1, p2, k3, k4, k5, k6, s1, s2, s3, s4, tau_x, tau_y
            }
            # if the dvec has more than 12 (and 8, 5, 4) elements, remove the last elements if they are all zeros
            # that means, the possible dvecs are: [k1, k2, p1, p2], [k1, k2, p1, p2, k3], [k1, k2, p1, p2, k3, ..., k6], 
            # [k1, k2, p1, p2, k3, ..., k6, s1, ... s4], and [k1, k2, p1, p2, k3, ..., k6, s1, ..., s4, tau_x, tau_y] 
            this_dvec = camera_parameters[cam_names[icam]]['dvec']
            if this_dvec.size > 4 and np.all(this_dvec[4:] == 0):
                this_dvec = this_dvec[:4]
            if this_dvec.size > 5 and np.all(this_dvec[5:] == 0):
                this_dvec = this_dvec[:5]
            if this_dvec.size > 8 and np.all(this_dvec[8:] == 0):
                this_dvec = this_dvec[:8]
            if this_dvec.size > 12 and np.all(this_dvec[12:] == 0):
                this_dvec = this_dvec[:12]
            camera_parameters[cam_names[icam]]['dvec'] = this_dvec
            # print date/time, and the camera parameters to the print_widget if provided
            if print_widget is not None:
                print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                print_widget.insert(tk.END, f"#   Camera '{cam_names[icam]}':\n")
                print_widget.insert(tk.END, f"#     image_size: {camera_parameters[cam_names[icam]]['image_size']}\n")
                print_widget.insert(tk.END, f"#     rvec: {camera_parameters[cam_names[icam]]['rvec'].flatten()}\n")
                print_widget.insert(tk.END, f"#     tvec: {camera_parameters[cam_names[icam]]['tvec'].flatten()}\n")
                print_widget.insert(tk.END, f"#     cmat:\n{camera_parameters[cam_names[icam]]['cmat']}\n")
                print_widget.insert(tk.END, f"#     dvec: {camera_parameters[cam_names[icam]]['dvec'].flatten()}\n")
        # poi_definition:
        # get the pois_definition worksheet
        # the first row is the headers, which are "poi_name", "xw", "yw", "zw", 
        # "xi__cam_name_1", "yi__cam_name_1", "x0__cam_name_1", "y0__cam_name_1", "w__cam_name_1", "h__cam_name_1", 
        # "xi__cam_name_2", "yi__cam_name_2", "x0__cam_name_2", "y0__cam_name_2", "w__cam_name_2", "h__cam_name_2", 
        # but some pois may only have xi and yi without x0, y0, w, h, for example, 
        # "xi__cam_name_3", "yi__cam_name_3",
        # "xi__cam_name_4", "yi__cam_name_4", and so on.
        # After reading data, the dictionary pois_definition will be like this:
        # pois_definition['poi_name_1']['Xw'] = np.array([xw, yw, zw]), the world coordinates of the point of interest 
        # pois_definttion['poi_name_1']['Xi']['cam_name_1'] = np.array([xi, yi]), the image coordinates of the point of interest in the camera image, 
        # if properties of x0, y0, w, h are available, they will be stored in the same dictionary as well.
        # For example, pois_definition['poi_name_1']['Tmplt']['cam_name_1'] = np.array([x0, y0, w, h], dtype=np.int32)
        # where x0, y0 are the top-left corner of the template, w is the width, and h is the height of the template.
        if 'pois_definition' not in wb.sheetnames:
            print("# Error: The project file does not contain a 'pois_definition' worksheet.")
            # if print_widget is provided, print date/time, and a message to it
            if print_widget is not None:
                print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                print_widget.insert(tk.END, "#   Error: The project file does not contain a 'pois_definition' worksheet.\n")
        else:
            ws_pois_definition = wb['pois_definition']
            pois_definition = {}
            headers = [cell.value for cell in ws_pois_definition[1]]
            # strip all strings in headers
            headers = [header.strip() for header in headers if isinstance(header, str)]
            # read the pois_definition worksheet
            for row in ws_pois_definition.iter_rows(min_row=2, values_only=True):
                poi_name = row[0].strip() if isinstance(row[0], str) else None
                if poi_name is None:
                    continue
                # create an empty dictionary for this poi, pois_definition[poi_name] 
                pois_definition[poi_name] = {}
                # get the world coordinates of the poi
                pois_definition[poi_name]['Xw'] = np.array(row[1:4], dtype=np.float64)
                # get the image coordinates of the poi in each camera
                pois_definition[poi_name]['Xi'] = {}
                # get the template of the poi in each camera
                pois_definition[poi_name]['Tmplt'] = {}
                for i in range(4, len(headers)):
                    cam_name = headers[i].split('__')[1]
                    if cam_name not in cam_names:
                        print(f"# Warning: Camera name '{cam_name}' in the header not found in camera_parameters. Skipping this camera.")
                        continue
                    # check if the property name is xi, yi, x0, y0, w, or h
                    prop_name = headers[i].split('__')[0].strip()
                    if prop_name not in ['xi', 'yi', 'x0', 'y0', 'w', 'h']:
                        print(f"# Warning: Property name '{prop_name}' in the header not recognized. Skipping this property.")
                        continue
                    # if prop_name is xi, read two elements (assuming yi follows xi)
                    if prop_name == 'xi':
                        xi = row[i]
                        yi = row[i + 1] if (i + 1) < len(row) else None
                        if yi is not None:
                            pois_definition[poi_name]['Xi'][cam_name] = np.array([xi, yi], dtype=np.float64)
                    elif prop_name == 'yi':
                        continue
                    # if prop_name is x0, read four elements (assuming y0, w, h follow x0)
                    if prop_name == 'x0':
                        x0 = row[i]
                        y0 = row[i + 1] if (i + 1) < len(row) else None
                        w = row[i + 2] if (i + 2) < len(row) else None
                        h = row[i + 3] if (i + 3) < len(row) else None
                        if y0 is not None and w is not None and h is not None:
                            pois_definition[poi_name]['Tmplt'][cam_name] = np.array([x0, y0, w, h], dtype=np.int32)
                # end of column 4 to end of this row 
            # end of this row 
            # if print_widget is provided, print date/time, and the names of all pois to it
            if print_widget is not None:
                print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                print_widget.insert(tk.END, "#   Points of Interest (POIs) defined in the project file:\n")
                for poi_name in pois_definition.keys():
                    print_widget.insert(tk.END, f"#     {poi_name}\n")
            # end of pois_definition worksheet
        # end of if 'pois_definition' in wb.sheetnames
        #
        # image_sources:
        # image_sources:
        # get the image_sources worksheet
        # For any file in the image_sources worksheet of this project file (.xlsx), 
        # if that path is a relative path, it is relative to the project file path (project_file_path).
        # For any file in the file-list text file, if that path is a relative path, it is relative to the file-list text file path.
        # For example, if the project file is /abs_path/to/project_file.xlsx,
        #   and the image source file path is cam_1/video1.mp4, 
        #   that means the video file is at /abs_path/to/cam_1/video1.mp4
        # For example, if the project file is /abs_path/to/project_file.xlsx,
        #   and the image source file path is cameras/file_list_camera3.txt,
        #   that means the file-list text file is at /abs_path/to/cameras/file_list_camera3.txt
        #   and if one of the file names in the file-list text file is cam_3/image1.jpg
        #   and the image file is at /abs_path/to/cameras/cam_3/image1.jpg
        if 'image_sources' not in wb.sheetnames:
            print("# Error: The project file does not contain a 'image_sources' worksheet.")
            # if print_widget is provided, print date/time, and a message to it
            if print_widget is not None:
                print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                print_widget.insert(tk.END, "#   Error: The project file does not contain a 'image_sources' worksheet.\n")
        else:
            ws_image_sources = wb['image_sources']
            image_sources = {}
            for row in ws_image_sources.iter_rows(min_row=2, values_only=True):
                camera_name = row[0].strip() if isinstance(row[0], str) else None
                image_source_file_path = row[1].strip() if isinstance(row[1], str) else None
                if camera_name is None or image_source_file_path is None:
                    continue
                # check if the image source file path is a relative path or an absolute path
                if not os.path.isabs(image_source_file_path):
                    # relative path, make it absolute by joining with the project file path
                    image_source_file_path = os.path.join(os.path.dirname(project_file_path), image_source_file_path)
                # check if the file exists
                if not os.path.isfile(image_source_file_path):
                    print(f"# Warning: The image source file {image_source_file_path} does not exist. You may want to check the path.")
                # check the file extension, if it is a .txt file, read the file and get the list of image files
                if image_source_file_path.endswith('.txt'):
                    # if this file exists, read the file and get the list of image files
                    if not os.path.isfile(image_source_file_path):
                        print(f"# Warning: The file-list text file {image_source_file_path} does not exist. You may want to check the path.")
                        continue
                    with open(image_source_file_path, 'r') as f:
                        image_files = [line.strip() for line in f.readlines() if line.strip()]
                    # convert to absolute paths if they are relative paths
                    image_files = [os.path.join(os.path.dirname(image_source_file_path), img) if not os.path.isabs(img) else img for img in image_files]
                    image_sources[camera_name] = image_files
                else:
                    # assume it is a video file, just store the file path
                    image_sources[camera_name] = image_source_file_path
                # print the image source file path to the print_widget if provided
                if print_widget is not None:
                    print_widget.insert(tk.END, f"# {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n")
                    print_widget.insert(tk.END, f"#   Camera '{camera_name}': Image source file path: {image_source_file_path}\n")
            # end of each row in image_sources worksheet
        # end of image_sources worksheet
   
    return basic_info, camera_parameters, pois_definition, image_sources
