"""
Cameras Class
=============
This class `Cameras` provides a convenient way to load, save, and visualize camera parameters stored in Excel or CSV files.
It is especially useful in computer vision applications where multiple calibrated cameras are used and their intrinsic
and extrinsic parameters must be handled.

How to Prepare the Input Excel/CSV File
---------------------------------------
* Prepare a file in `.xlsx` or `.csv` format.
* The first column must be named `parameters`, listing each camera parameter in the following fixed order:

    image_width
    image_height
    rvec_x
    rvec_y
    rvec_z
    tvec_x
    tvec_y
    tvec_z
    cmat11_fx
    cmat12
    cmat13_cx
    cmat21
    cmat22_fy
    cmat23_cy
    cmat31
    cmat32
    cmat33
    k1
    k2
    p1 (optional)
    p2 (optional)
    k3 (optional)
    k4 (optional)
    k5 (optional)
    k6 (optional)
    s1 (optional)
    s2 (optional)
    s3 (optional)
    s4 (optional)
    tau_x (optional)
    tau_y (optional)

* The following columns represent each camera, using user-defined camera names.

Example layout in Excel/CSV:

| parameters   | camA         | camB         | camC         |
|--------------|--------------|--------------|--------------|
| image_width  | 2704         | 3840         | 3840         |
| image_height | 2028         | 3360         | 3360         |
| rvec_x       | 0.37         | 0.35         | 0.62         |
| rvec_y       | -0.34        | -0.38        | -0.0077      |
| rvec_z       | 0.12         | 0.11         | 0.10         |
| ...          | ...          | ...          | ...          |

* If a parameter is not provided, it defaults to 0.0 (only applies to optional parameters).
* The worksheet must be named `camera_parameters` unless the file only contains one worksheet.

Mathematical Representation
----------------------------
* rvec and tvec: 3D rotation and translation vectors. Used to represent extrinsic transformations.
* cmat: 3x3 intrinsic matrix:
      [[fx, 0, cx],
       [0, fy, cy],
       [0,  0,  1]]
* dvec: Distortion coefficients in the order:
      [k1, k2, p1, p2, k3, k4, k5, k6, s1, s2, s3, s4, tau_x, tau_y]
   Only the non-zero tail is kept for space efficiency.

Class Methods
-------------

__init__(self):
    Initializes an empty Cameras object with no loaded data.

load_from_file(self, cameras_file=None):
    Loads camera parameters from an Excel or CSV file.
    If no filename is provided, a file dialog will pop up.
    Returns None on error or cancel.

    Example:
        cams = Cameras()
        cams.load_from_file("my_cameras.xlsx")

save_to_file(self, cameras_file=None):
    Saves current camera parameters to an Excel or CSV file.
    If no filename is provided, a file dialog will pop up.
    Only the worksheet "camera_parameters" is affected in Excel.

    Example:
        cams.save_to_file("output.xlsx")

plot_cameras(self):
    Calls `plot_cameras()` function from the module `plotCameras3d`.
    Displays a 3D visualization of all loaded camera poses and orientations.

    Example:
        cams.plot_cameras()

__str__(self) and __repr__(self):
    Returns a string summary of the object, listing number of cameras and their names.

    Example:
        print(cams)

Usage Example
-------------
from cameras import Cameras

# Create a Cameras object
cams = Cameras()

# Load camera data from a file
cams.load_from_file("my_cameras.xlsx")

# Print summary
print(cams)

# Visualize in 3D
cams.plot_cameras()

# Save to another file
cams.save_to_file("output_cameras.xlsx")
"""

import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import numpy as np
import csv
from plotCameras3d import plot_cameras

class Cameras:
    def __init__(self):
        self.camera_parameters = {}

    def load_from_file(self, cameras_file=None):
        if not cameras_file:
            tk.Tk().withdraw()
            cameras_file = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Text files", "*.txt")]
            )
            if not cameras_file:
                return None

        ext = os.path.splitext(cameras_file)[1].lower()
        if ext == ".xlsx":
            wb = load_workbook(cameras_file, data_only=True)
            if len(wb.sheetnames) == 1:
                ws = wb.active
            elif "camera_parameters" in wb.sheetnames:
                ws = wb["camera_parameters"]
            else:
                print("Error: Worksheet 'camera_parameters' not found.")
                return None
            data = [[cell.value for cell in row] for row in ws.iter_rows()]

        elif ext in [".csv", ".txt"]:
            with open(cameras_file, newline='') as csvfile:
                data = list(csv.reader(csvfile))
        else:
            print("Error: Unsupported file type.")
            return None

        param_names = [row[0] for row in data if row[0] is not None]
        all_params = ["image_width", "image_height", "rvec_x", "rvec_y", "rvec_z", "tvec_x", "tvec_y", "tvec_z",
                      "cmat11_fx", "cmat12", "cmat13_cx", "cmat21", "cmat22_fy", "cmat23_cy",
                      "cmat31", "cmat32", "cmat33", "k1", "k2", "p1", "p2", "k3", "k4", "k5", "k6",
                      "s1", "s2", "s3", "s4", "tau_x", "tau_y"]

        param_indices = {name: i for i, name in enumerate(param_names) if name in all_params}
        cam_names = data[0][1:]
        for j, cam in enumerate(cam_names):
            values = {p: float(data[i][j+1]) if j+1 < len(data[i]) and data[i][j+1] else 0.0 for p, i in param_indices.items()}
            img_size = (int(values['image_width']), int(values['image_height']))
            rvec = np.array([values['rvec_x'], values['rvec_y'], values['rvec_z']]).reshape(3,1)
            tvec = np.array([values['tvec_x'], values['tvec_y'], values['tvec_z']]).reshape(3,1)
            cmat = np.array([
                [values['cmat11_fx'], values['cmat12'], values['cmat13_cx']],
                [values['cmat21'], values['cmat22_fy'], values['cmat23_cy']],
                [values['cmat31'], values['cmat32'], values['cmat33']]
            ])
            dvec_list = [values.get(k, 0.0) for k in ["k1", "k2", "p1", "p2", "k3", "k4", "k5", "k6",
                                                      "s1", "s2", "s3", "s4", "tau_x", "tau_y"]]
            while dvec_list and dvec_list[-1] == 0.0:
                dvec_list.pop()
            dvec = np.array(dvec_list).reshape(-1,1)

            self.camera_parameters[cam] = {
                'image_size': img_size,
                'rvec': rvec,
                'tvec': tvec,
                'cmat': cmat,
                'dvec': dvec
            }

    def save_to_file(self, cameras_file=None):
        if not cameras_file:
            tk.Tk().withdraw()
            cameras_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Text files", "*.txt")]
            )
            if not cameras_file:
                return None

        ext = os.path.splitext(cameras_file)[1].lower()
        header = ["parameters"] + list(self.camera_parameters.keys())
        param_order = ["image_width", "image_height", "rvec_x", "rvec_y", "rvec_z", "tvec_x", "tvec_y", "tvec_z",
                       "cmat11_fx", "cmat12", "cmat13_cx", "cmat21", "cmat22_fy", "cmat23_cy",
                       "cmat31", "cmat32", "cmat33", "k1", "k2", "p1", "p2", "k3", "k4", "k5", "k6",
                       "s1", "s2", "s3", "s4", "tau_x", "tau_y"]

        rows = []
        for param in param_order:
            row = [param]
            for cam in self.camera_parameters:
                p = self.camera_parameters[cam]
                value = None
                if param == 'image_width': value = p['image_size'][0]
                elif param == 'image_height': value = p['image_size'][1]
                elif param.startswith("rvec_"): value = p['rvec']["xyz".index(param[-1])][0]
                elif param.startswith("tvec_"): value = p['tvec']["xyz".index(param[-1])][0]
                elif param.startswith("cmat"):
                    idx = param_order.index(param) - param_order.index("cmat11_fx")
                    value = p['cmat'].flatten()[idx]
                else:
                    dvec_idx = ["k1", "k2", "p1", "p2", "k3", "k4", "k5", "k6", "s1", "s2", "s3", "s4", "tau_x", "tau_y"].index(param)
                    if dvec_idx < p['dvec'].shape[0]:
                        value = p['dvec'][dvec_idx][0]
                    else:
                        value = 0.0
                row.append(value)
            rows.append(row)

        if ext == ".xlsx":
            if os.path.exists(cameras_file):
                wb = load_workbook(cameras_file)
            else:
                wb = Workbook()
            if "camera_parameters" in wb.sheetnames:
                del wb["camera_parameters"]
            ws = wb.create_sheet("camera_parameters")
            ws.append(header)
            for row in rows:
                ws.append(row)
            wb.save(cameras_file)

        else:
            with open(cameras_file, "w", newline='') as f:
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(rows)

    def plot_cameras(self):
        from plotCameras3d import plot_cameras, plotCameras3d
        rvecs = [self.camera_parameters[c]['rvec'] for c in self.camera_parameters]
        tvecs = [self.camera_parameters[c]['tvec'] for c in self.camera_parameters]
        cmats = [self.camera_parameters[c]['cmat'] for c in self.camera_parameters]
        dvecs = [self.camera_parameters[c]['dvec'] for c in self.camera_parameters]
        img_sizes = [self.camera_parameters[c]['image_size'] for c in self.camera_parameters]
        plotCameras3d(img_sizes, rvecs, tvecs, cmats, dvecs)
#        plot_cameras(rvecs, tvecs, cmats, dvecs, img_sizes)

    def __str__(self):
        return f"Cameras({len(self.camera_parameters)} cameras: {list(self.camera_parameters.keys())})"

    def __repr__(self):
        return self.__str__()

# Example usage:
if __name__ == '__main__':
    cams = Cameras()
    cams.load_from_file()   # Will prompt file dialog if no argument
    print(cams)
    cams.plot_cameras()
    cams.save_to_file()     # Will prompt file dialog if no argument
