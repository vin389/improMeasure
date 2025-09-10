import os, sys
import time, datetime
import tkinter as tk
import tkinter.filedialog
import tkinter.messagebox
#import tkinter.scrolledtext
import numpy as np
import cv2 as cv

from improCalib import newArrayByMapping, countCalibPoints
from improMisc import uigetfile, uiputfile, uigetfiles_tupleFullpath
from improStrings import npFromString, stringFromNp, npFromTupleNp
from drawPoints import drawPoints
from imshow2 import imshow2
from writeCamera import writeCamera
from draw3dMesh import draw3dMesh
from chessboard_object_points import chessboard_object_points

def tkCalib_printMessage(msg: str):
    strfNow = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("%s: %s" % (strfNow, msg))
    return

def tkCalib():
#    camParams = np.zeros(14, dtype=np.float64)
    imgSize = np.zeros(2, dtype=int)
    rvec = np.zeros((3, 1), dtype=float)
    tvec = np.zeros((3, 1), dtype=float)
    cmat = np.eye(3, dtype=float)
    dvec = np.zeros((1, 5), dtype=float)
    # constants
    strFlags = ['CALIB_USE_INTRINSIC_GUESS', 'CALIB_FIX_ASPECT_RATIO',
        'CALIB_FIX_PRINCIPAL_POINT', 'CALIB_ZERO_TANGENT_DIST',
        'CALIB_FIX_FOCAL_LENGTH', 'CALIB_FIX_K1',
        'CALIB_FIX_K2', 'CALIB_FIX_K3', 'CALIB_FIX_K4',
        'CALIB_FIX_K5', 'CALIB_FIX_K6', 'CALIB_RATIONAL_MODEL',
        'CALIB_THIN_PRISM_MODEL', 'CALIB_FIX_S1_S2_S3_S4',
        'CALIB_TILTED_MODEL', 'CALIB_FIX_TAUX_TAUY', 'CALIB_USE_QR',
        'CALIB_FIX_TANGENT_DIST', 'CALIB_FIX_INTRINSIC',
        'CALIB_SAME_FOCAL_LENGTH', 'CALIB_ZERO_DISPARITY',
        'CALIB_USE_LU', 'CALIB_USE_EXTRINSIC_GUESS' ]
    #
    global win
    win = tk.Tk()
    win.geometry("1600x900")
    frame1 = tk.Frame(win)
    frame1.pack(side=tk.LEFT, anchor='n')
    frame2 = tk.Frame(win)
    frame2.pack(side=tk.LEFT, anchor='n')
    frame3 = tk.Frame(win)
    frame3.pack(side=tk.LEFT, anchor='n')
    frame4 = tk.Frame(win)
    frame4.pack(side=tk.LEFT, anchor='n')

    # #######################################################
    # a pair of ratio buttons of [1-image calib] and [chessboard calib]
    # #######################################################
    # define function of radio button
    # if the ratio button of [1-image calib] is clicked,
    # it enables 'Calibrate camera (1-photo)' button 
    # and disables 'Find chessboard corners' button and 'Calibrate camera (chessboard)' button
    # if the ratio button of [chessboard calib] is clicked,
    # it enables 'Find chessboard corners' button and 'Calibrate camera (chessboard)' button
    # and disables 'Calibrate camera (1-photo)' button
    # Note. The btCalib, btFindCbCorners, and btCalibCb are defined in the following sections
    #       after the radio buttons.
    def rb_clicked():
        if rbVar.get() == 0:
            btCalib.configure(state='normal')
            btFindCbCorners.configure(state='disabled')
            edCbParams.config(state=tk.DISABLED, bg="#F0F0F0")
            btCalibCb.configure(state='disabled')
            edGridImgIdx.configure(state='disabled')
            pass
        else:
            btCalib.configure(state='disabled')
            btFindCbCorners.configure(state='normal')
            edCbParams.config(state=tk.NORMAL, bg="white")
            btCalibCb.configure(state='normal')
            edGridImgIdx.configure(state='normal')
            pass
        return
    # create a pair of radio buttons
    rbVar = tk.IntVar()
    rb1 = tk.Radiobutton(frame1, text='single-image calib', variable=rbVar, value=0, command=rb_clicked)
    # align the radio buttons to the left side of the window
    rb1.pack(anchor=tk.W)
    rb2 = tk.Radiobutton(frame1, text='chessboard calib', variable=rbVar, value=1, command=rb_clicked)
    rb2.pack(anchor=tk.W)
    # set initial value of radio button
    rbVar.set(0)

    # #######################################################
    # Button and edit text of [3D coordinates (World coord.)]
    # #######################################################
    # button of 3D coordinates (World coord.)
    btCoord3d = tk.Button(frame1, height=1, text='3D coordinates (World coord.)')
    btCoord3d.pack()
    # edit text of 3D coordinates
    edCoord3d = tk.Text(frame1, width=40, height=3, undo=True, 
                        autoseparators=True, maxundo=-1)
    edCoord3d.pack()
    #   set initial text for demonstration
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_coord3d.npy'))
        theStr = stringFromNp(theMat, ftype='txtf', sep='\t')
        edCoord3d.delete(0., tk.END)
        edCoord3d.insert(0., theStr)
    except:
        tkCalib_printMessage("# Warning: Failed to load tkCalib_init_coord3d.npy")
        edCoord3d.delete(0., tk.END)
        edCoord3d.insert(0., ' 0 0 0 \n 1 0 0 \n 1 1 0 \n 0 1 0 \n nan nan nan')  
    # button command function 
    def btCoord3d_clicked():
        strPoints3d = edCoord3d.get(0., tk.END)
        try:
            points3d = npFromString(strPoints3d).reshape((-1, 3))
        except:
            tkCalib_printMessage('Cannot parse 3D points (size: %d)' % 
                                 (int(points3d.size)))
        strPointsImg= edCoordImg.get(0., tk.END)
        try:
            points2d = npFromString(strPointsImg).reshape((-1, 2))
        except:
            tkCalib_printMessage('Cannot parse 2D points (size: %d)' % 
                                 (int(points2d.size)))
        validsOfPoints3d, validsOfPoints2d, validCalibPoints, \
            idxAllToValid, idxValidToAll, validPoints3d, validPoints2d = \
            countCalibPoints(points3d, points2d)
        npts3d = points3d.shape[0]
        nval3d = np.sum(validsOfPoints3d)
        npts2d = points2d.shape[0]
        nval2d = np.sum(validsOfPoints2d)
        nValidCalibPoints = np.sum(validCalibPoints)
        # print info
        tkCalib_printMessage('Numbers of valid/total 3D points: %d/%d.' % 
                             (nval3d, npts3d))
        tkCalib_printMessage('Numbers of valid/total 2D points: %d/%d.' % 
                             (nval2d, npts2d))
        if (npts3d == npts2d):
            tkCalib_printMessage('Number of valid calibration points: %d' % 
                                 nValidCalibPoints)
            strValidPoints = '  Valid calib points (index is 0 based):'
            for i in range(npts3d):
                if validCalibPoints[i] == 1:
                    strValidPoints += ' %d' % i
            tkCalib_printMessage(strValidPoints)
        return
    # set button command function 
    btCoord3d.configure(command=btCoord3d_clicked)
    # #######################################################
    # Button and edit text of [Image 2D coordinates]
    # #######################################################
    # button of image coordinates
    btCoordImg = tk.Button(frame1, text='Image 2D coordinates')
    btCoordImg.pack()
    # edit text of image coordinates
    edCoordImg = tk.Text(frame1, width=40, height=3, undo=True, 
                         autoseparators=True, maxundo=-1)
    edCoordImg.pack()
    #   set initial text for demonstration
    edCoordImg.delete(0., tk.END)
    #   set initial text for demonstration
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_coordImg.npy'))
        theStr = stringFromNp(theMat, sep='\t')
        edCoordImg.insert(0., theStr)
    except:
        tkCalib_printMessage("# Warning: Failed to load tkCalib_init_coordImg.npy")
        edCoordImg.insert(0., '1. 1. \n 2. 2. \nnan nan \n 1. 2. \n 3. 3.')
    # button command function
    def btCoordImg_clicked():
        strPoints3d = edCoord3d.get(0., tk.END)
        strPointsImg= edCoordImg.get(0., tk.END)
        try:
            points3d = npFromString(strPoints3d).reshape((-1, 3))
        except:
            tkCalib_printMessage('# Error: Cannot parse 3D points (size: %d)' % 
                                 (int(points3d.size)))
        try:
            points2d = npFromString(strPointsImg).reshape((-1, 2))
        except:
            tkCalib_printMessage('# Error: Cannot parse 2D points (size: %d)' % 
                                 (int(points2d.size)))
        validsOfPoints3d, validsOfPoints2d, validCalibPoints, \
            idxAllToValid, idxValidToAll, validPoints3d, validPoints2d = \
            countCalibPoints(points3d, points2d)
        npts3d = points3d.shape[0]
        nval3d = np.sum(validsOfPoints3d)
        npts2d = points2d.shape[0]
        nval2d = np.sum(validsOfPoints2d)
        nValidCalibPoints = np.sum(validCalibPoints)
        # print info
        tkCalib_printMessage('Numbers of valid/total 3D points: %d/%d.' % 
                             (nval3d, npts3d))
        tkCalib_printMessage('Numbers of valid/total 2D points: %d/%d.' % 
                             (nval2d, npts2d))
        if (npts3d == npts2d):
            tkCalib_printMessage('Number of valid calibration points: %d' % 
                                 nValidCalibPoints)
            strValidPoints = '  Valid calib points (index is 0 based):'
            for i in range(npts3d):
                if validCalibPoints[i] == 1:
                    strValidPoints += ' %d' % i
            tkCalib_printMessage(strValidPoints)
        return
    # set command     
    btCoordImg.configure(command=btCoordImg_clicked)  
    # #######################################################
    # Button of [Select calibration image(s) ...]
    # #######################################################
    btFile = tk.Button(frame1, text='Select calibration image(s) ...')
    btFile.pack()
    edFile = tk.Text(frame1, width=40, height=2, undo=True, 
                     autoseparators=True, maxundo=-1)
    edFile.pack()
    #   set initial text for demonstration
    try:
        tfile = open(os.path.join(os.getcwd(), 'tkCalib_init_imgFile.txt'), "r")
        fname = tfile.read()
        tfile.close()
        edFile.delete(0., tk.END)
        edFile.insert(0., fname)
    except:
        edFile.insert(0., 'c:/images/calibration.bmp')
        
    #   btFile 'Select image ...'
    def btFile_clicked():
        # select file
        fpath = edFile.get(0., tk.END)
        try:
            # fpath could be 1 file or multiple files
            fpath = fpath.split()[0]
        except:
            pass
        initDir = os.path.split(fpath)[0]
        # get file(s) from file dialog (format: tuple of full paths)
        uFullpaths = uigetfiles_tupleFullpath(initialDirectory=initDir)
        # set edFile text to the full path of select file(s)
        # if user selects 1 file, edFile text would be the full path of the file
        # if user selects multiple files, edFile text would be full paths of all files and are separated with \n
        # the image size would be set to that of the first file (index [0])
        if len(uFullpaths) > 1:
            edFile.delete(0., tk.END)
            edFile.insert(0., '\n'.join(uFullpaths))
        else:
            edFile.delete(0., tk.END)
            edFile.insert(0., uFullpaths[0])
        btImgSize_clicked() 
        return
    # set button command function 
    btFile.configure(command=btFile_clicked)
            
    # #######################################################
    # Button of [Image size:] (width height)
    # #######################################################
    btImgSize = tk.Button(frame1, text='Image size: width height')
    btImgSize.pack()
    edImgSize = tk.Text(frame1, width=40, height=1, undo=True, 
                        autoseparators=True, maxundo=-1)
    edImgSize.pack()
    #   set initial text for demonstration
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_imgSize.npy')).astype(int)
        theStr = stringFromNp(theMat, sep='\t')
        edImgSize.delete(0., tk.END)
        edImgSize.insert(0., theStr)
    except:
        tkCalib_printMessage("# Warning: Cannot parse tkCalib_init_imgSize.npy")
        edImgSize.delete(0., tk.END)
        edImgSize.insert(0., '1920 1080')
    # define functions
    def btImgSize_clicked():
        # try to open the image file and check the image size
        try:
            fname = edFile.get(0., tk.END)
            # if fname contains multiple files with delimiter of \n \t or space,
            # we only take the first one.
            fname = fname.split()[0]
            img = cv.imread(fname)
            if type(img) == np.ndarray and img.shape[0] >= 1:
                imgH = img.shape[0]
                imgW = img.shape[1]
                edImgSize.delete(0., tk.END)
                edImgSize.insert(0., '%d %d' % (imgW, imgH))
            # update initial guess of cmat
            theStr = edCmatGuess.get(0., tk.END)
            theMat = npFromString(theStr).reshape((3,3))
            if theMat.shape[0] == 3 and theMat.shape[1] == 3:
                theMat[0, 2] = imgW / 2.0 - 0.5
                theMat[1, 2] = imgH / 2.0 - 0.5
                theStr = stringFromNp(theMat, sep='\t')
                edCmatGuess.delete(0., tk.END)
                edCmatGuess.insert(0., theStr)
        except:
            tk.messagebox.showerror(title="Error", message="Cannot read the image file.")
            print("Cannot read the image file.")
        strImgSize = edImgSize.get(0., tk.END)
        print('Image size: ', strImgSize)
        return
    # imgSizeFromBt() determines image size. 
    # imgSizeFromBt() is called by btCalib_clicked(), btDrawPoints_clicked(), 
    #   btDrawPointsUndistort_clicked(), 
    def imgSizeFromBt():
        strImgSize = edImgSize.get(0., tk.END)
        imgSize = npFromString(strImgSize)
        if type(imgSize) != np.ndarray or imgSize.size != 2:
            errMsg = '# error: Image size is invalid (edit text is %s)' % (strImgSize)
            tk.messagebox.showerror(title="Error", message="errMsg")
            return None
        return imgSize.astype(int).flatten()
    # set button command function
    btImgSize.configure(command=btImgSize_clicked)
    
    # #######################################################
    # Button of [Camera mat (init guess)]
    # #######################################################
    btCmatGuess = tk.Button(frame1, text='Camera mat (init guess)')
    btCmatGuess.pack()
    edCmatGuess = tk.Text(frame1, width=40, height=2, undo=True, 
                          autoseparators=True, maxundo=-1)
    edCmatGuess.pack()
    #   set initial text for demonstration
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_cmatGuess.npy'))
        theStr = stringFromNp(theMat, sep='\t')
        edCmatGuess.delete(0., tk.END)
        edCmatGuess.insert(0., theStr)
    except:
        print("Warning: Cannot load tkCalib_init_cmatGuess.npy")
        edCmatGuess.delete(0., tk.END)
#        edCmatGuess.insert(0., ' 5000. 0 0 \n 0 5000. 0 \n 1999.5 1999.5 1')  # This is wrong. Incorrectly written in transposed way.
        edCmatGuess.insert(0., ' 5000. 0 1999.5 \n 0 5000. 1999.5 \n 0 0 1')  # revised on 2024/08/28
    #   button 'Camera mat (init guess)' as a label
    # defind button command function 
    def btCmatGuess_clicked():
        try:
            strCmatGuess = edCmatGuess.get(0., tk.END)
            cmatGuess = npFromString(strCmatGuess).reshape((3,3))
        except:
            cmatGuess = np.array([])
        print('Initial guess of camera matrix:\n', cmatGuess)
        return cmatGuess
    # set button command function 
    btCmatGuess.configure(command=btCmatGuess_clicked)
    # #######################################################
    # Button of [Distortion coeff. (init guess)]
    # #######################################################
    btDvecGuess = tk.Button(frame1, text='Distortion coeff. (init guess)')
    btDvecGuess.pack()
    edDvecGuess = tk.Text(frame1, width=40, height=2, undo=True, 
                          autoseparators=True, maxundo=-1)
    edDvecGuess.pack()
    #   set initial text for demonstration
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_dvecGuess.npy'))
        theStr = stringFromNp(theMat, sep='\t')
        edDvecGuess.delete(0., tk.END)
        edDvecGuess.insert(0., theStr)
    except:
        print("Warning: Cannot load tkCalib_init_dvecGuess.npy")
        edDvecGuess.delete(0., tk.END)
        edDvecGuess.insert(0., '0. 0. 0. 0.')
    # define button command function 
    def btDvecGuess_clicked():
        strDvecGuess = edDvecGuess.get(0., tk.END)
        try:
            cdvecGuess = npFromString(strDvecGuess).reshape((1,-1))
        except:
            cdvecGuess = np.array([])
        print('Initial guess of distortion vector:\n', cdvecGuess)
        return cdvecGuess
    # set button command function 
    btDvecGuess.configure(command=btDvecGuess_clicked)

    # #######################################################
    # Button of [Find chessboard corners]
    # #######################################################
    btFindCbCorners = tk.Button(frame1, text='Find chessboard corners')
    btFindCbCorners.pack()
    edCbParams = tk.Text(frame1, width=40, height=1, undo=True, 
                          autoseparators=True, maxundo=-1)
    edCbParams.pack()
    #   set initial text for demonstration
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_cboardParams.npy'))
#        theStr = stringFromNp(theMat, sep='\t')
        theStr = '%d %d %f %f' % (theMat[0], theMat[1], theMat[2], theMat[3])
        edCbParams.delete(0., tk.END)
        edCbParams.insert(0., theStr)
    except:
        print("Warning: Cannot load tkCalib_init_cboardParams.npy")
        edCbParams.delete(0., tk.END)
        edCbParams.insert(0., '7 7 25.4 25.4 ')
    # define button command function
    def btFindCbCorners_clicked():
        # get chessboard files
        try:
            cbFiles = edFile.get(0., tk.END).split()
            nCbFiles = len(cbFiles)
        except:
            tk.messagebox.showerror(title="Error", message="# Error: Cannot get calibration files.")
            return 
        # get chessboard parameters (nCornersX, nCornersY, dxCorners, dyCorners)
        try:
            strCbParams = edCbParams.get(0., tk.END)
            npCbParams = npFromString(strCbParams)
            nCornersX = int(npCbParams[0])
            nCornersY = int(npCbParams[1])
            dxCorners = float(npCbParams[2])
            dyCorners = float(npCbParams[3])
        except:
            tk.messagebox.showerror(title="Error", message="# Error: Cannot get chessboard information.")
            return 
        # calibration flags
        try:
#            calibFlags = int(edFlags.get())
            calibFlags = ck_clicked()
        except:
            tk.messagebox.showerror(title="Error", message="# Error: Cannot get flags.")
            return 
        # allocate arrays (?)
        nFiles = len(cbFiles)
        # generate object points of chessboard corners
        corners_object_points_one_picture = chessboard_object_points(nCornersX, nCornersY, dxCorners, dyCorners)
        # clone the object points for all images. Stack vertically, make it 3D, i.e., nCbFiles * (nCornersX * nCornersY) * 3
        corners_object_points = np.tile(corners_object_points_one_picture, (nCbFiles, 1, 1))
        # print it to text box
        theStr = stringFromNp(corners_object_points.reshape(-1,3), ftype='txtf', sep='\t')
        edCoord3d.delete(0., tk.END)
        edCoord3d.insert(0., theStr)
        # try to find corners from the images
        try:
            imgPointsList = []
            cornerFoundPhotoIndices = []
            for icb in range(nFiles):
                fname = cbFiles[icb]
                img = cv.imread(fname, cv.IMREAD_GRAYSCALE)
                flags=cv.CALIB_CB_ADAPTIVE_THRESH + cv.CALIB_CB_NORMALIZE_IMAGE
                # The variable "found" is a boolean variable that indicates whether the corners are found or not.
                # If "found", the variable ptsThis will be a float32 (nCornersX * nCornersY) * 1 * 2 array (3D).
                # (If not "found", the variable ptsThis will be a None.)
                found, ptsThis = cv.findChessboardCorners(img, (nCornersX, nCornersY), flags)
                if found:
                    print('Found corners in %s' % (os.path.basename(fname)))
                    imgPointsList.append(ptsThis)
                    cornerFoundPhotoIndices.append(icb)
                else:
                    errMsg = '# Failed to find corners in %s' % (os.path.basename(fname))
                    tkCalib_printMessage(errMsg)
                    tk.messagebox.showerror(title='Warning', message=errMsg)
        except:
            errMsg = '# Failed to find corners'
            tk.messagebox.showerror(title='Error', message=errMsg)
            return
        if len(imgPointsList) <= 0:
            errMsg = '# Failed to find corners for all images.'
            tk.messagebox.showerror(title='Error', message=errMsg)
            return
        # 
        nFound = len(imgPointsList)
        # change the edFile text to the files that corners are successfully found
        theStr = '\n'.join([cbFiles[i] for i in cornerFoundPhotoIndices])
        edFile.delete(0., tk.END)
        edFile.insert(0., theStr)
        # convert imgPointsList to imgPoints2f
#        imgPoints2f = np.array(imgPointsList).reshape(nFound, -1, 2)
        imgPoints2f = np.array(imgPointsList).reshape(-1, 2)
        # print detected corners to 
        theStr = stringFromNp(imgPoints2f, sep='\t')
        edCoordImg.delete(0., tk.END)
        edCoordImg.insert(0., theStr)
        print(cbFiles, nCbFiles)
        strCbParams = edCbParams.get(0., tk.END)
        pass
    # set button command function 
    btFindCbCorners.configure(command=btFindCbCorners_clicked)

    # #######################################################
    # Frame 2
    # #######################################################
    # Checkbuttons of calibration flags
    # #######################################################
    # Checkbuttons of calibration flags 
    #   allowing user to switch on/off of every flag
    #   once clicked, sum of flags is displaced 
    # init of flags 
    edFlags = tk.Entry(frame2)
    edFlags.pack(anchor=tk.W)
    edFlags.delete(0, tk.END)
    edFlags.insert(0, 'Calib flags: 0')
    edFlags.config(state= "disabled")
    ckFlags = []
    ckValues = []
    # checkbox function
    def ck_clicked():
        # calculate sum of flags
        flags = 0
        for i in range(len(strFlags)):
            if ckValues[i].get() == 1:
                flags += eval('cv.' + strFlags[i])
        # display sum of flags
        #   set edFlags text to Calib flags: %d
        edFlags.config(state= "normal")
        edFlags.delete(0, tk.END)
        edFlags.insert(0, 'Calib flags: %d' % (flags))
        edFlags.config(state= "disabled")
        return flags
    for i in range(len(strFlags)):
        # generate statement string for creating checkbutton
        #   E.g., for i == 0
        #   "tk.Checkbutton(win, text='CALIB_USE_INTRINSIC_GUESS (0)',
        #                   command=ck_clicked, variable=ckValues[i])"
        evalStr = "tk.Checkbutton(frame2, text='"
        evalStr += strFlags[i] + " (%d)" % (eval('cv.' + strFlags[i])) + "', "
        evalStr += "command=ck_clicked, "
        evalStr += "variable=ckValues[%d]" % (i) + ")"
        # create a tk.IntVar() for checkbutton value
        ckValues.append(tk.IntVar())
        # create a checkbutton
        ckFlags.append(eval(evalStr)) # 
        # position a checkbutton
        ckFlags[i].pack(anchor=tk.W)
    # disable some flag checks that is not supported in tkCalib
    # as tkCalib assumes the distortion coefficients are k1, k2, p1, p2, ...
    # rather than s1, s2, .... 
    ckFlags[12].config(state= "disabled")
    ckFlags[13].config(state= "disabled")
    ckFlags[14].config(state= "disabled")
    ckFlags[15].config(state= "disabled")
    ckFlags[16].config(state= "disabled")
#    ckFlags[17].config(state= "disabled")
#    ckFlags[18].config(state= "disabled")
#    ckFlags[19].config(state= "disabled")
#    ckFlags[20].config(state= "disabled")
    ckFlags[21].config(state= "disabled")
    ckFlags[22].config(state= "disabled")

    # set initial flags (edFlags)
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_calibFlags.npy'))
        theInt = int(theMat)
        theStr = '%d' % theInt
        edFlags.delete(0, tk.END)
        edFlags.insert(0, theStr)
        for i in range(len(strFlags)):
            flagInt = eval('cv.' + strFlags[i])
            if theInt & flagInt > 0:
                ckValues[i].set(1)
            else:
                ckValues[i].set(0)
    except:
        tkCalib_printMessage('Cannot load calibration flags from file. Set default flags.')
        ckValues[0].set(1)  # to use intrinsic guess
        ckValues[2].set(1)  # to fix principal point
        ckValues[3].set(1)  # to zero tangent dist
        ckValues[6].set(1)  # to fix k2
        ckValues[7].set(1)  # to fix k3
        ckValues[8].set(1)  # to fix k4
        ckValues[9].set(1)  # to fix k5
        ckValues[10].set(1) # to fix k6
    ck_clicked()

    # #######################################################
    # Frame 3
    # #######################################################
    # Button of [Calibrate camera (1-photo)]
    # #######################################################
    btCalib = tk.Button(frame3, text='Calibrate camera (1-photo)', width=40, height=2)
    btCalib.pack()
    # define button command function 
    def btCalib_clicked():
        print("# *** Button [Calibrate camera] clicked:\\")
        # get object points and image points as Numpy array format
        # from text boxes (edCoord3d and edCoordImg)
        strPoints3d = edCoord3d.get(0., tk.END)
        strPoints2d = edCoordImg.get(0., tk.END)
        try:
            points3d = npFromString(strPoints3d).reshape((-1, 3))
        except:
            tkCalib_printMessage('Cannot parse 3D points (size: %d)' % 
                                 (int(points3d.size)))
        try:
            points2d = npFromString(strPoints2d).reshape((-1, 2))
        except:
            tkCalib_printMessage('Cannot parse 2D points (size: %d)' % 
                                 (int(points2d.size)))
        # remove nan points, get intersection set of non-nan points
        #　get the mapping
        validsOfPoints3d, validsOfPoints2d, validCalibPoints, \
            idxAllToValid, idxValidToAll, validPoints3d, validPoints2d = \
            countCalibPoints(points3d, points2d)
        # variables and parameters to run camera calibration 
        nCalPoints = validPoints3d.shape[0]
        objPoints = validPoints3d.reshape((1, nCalPoints, 3)).astype(np.float32)
        imgPoints = validPoints2d.reshape((1, nCalPoints, 2)).astype(np.float32)
        imgSize = imgSizeFromBt().astype(int).reshape(-1)
        cmat = npFromString(edCmatGuess.get(0., tk.END)).reshape((3,3))
        dvec = npFromString(edDvecGuess.get(0., tk.END)).reshape((1,-1))
        flags = ck_clicked()
        rvecs = np.array([])
        tvecs = np.array([])
        print("Calibrating camera based on %d 3D points." % objPoints.shape[1])
        print("Calibrating camera based on %d img points." % imgPoints.shape[1])
        # run calibration 
        ret, cmat, dvec, rvecs, tvecs = cv.calibrateCamera(
            objPoints, imgPoints, imgSize, cmat, dvec, rvecs, tvecs, flags) 
        # reshape the results (rvecs and tvecs to 3d arrays)
#        rvecs = npFromTupleNp(rvecs[0]).reshape((3,1))
#        tvecs = npFromTupleNp(tvecs[0]).reshape((3,1))
        # display result to text boxes
        edRvecs.delete(0., tk.END)
        edRvecs.insert(0., stringFromNp(rvecs[0].reshape(-1,1), sep='\t'))
        edTvecs.delete(0., tk.END)
        edTvecs.insert(0., stringFromNp(tvecs[0].reshape(-1,1), sep='\t'))
        edCmat.delete(0., tk.END)
        edCmat.insert(0., stringFromNp(cmat.reshape(3,3), sep='\t'))
        edDvec.delete(0., tk.END)
        edDvec.insert(0., stringFromNp(dvec.reshape(-1,1), sep='\t'))
        # calculate camera position
        #   rvecs and tvecs would be a tuple of 3x1 float array(s)
#        campos = tvecs.copy() # make campos the same type and shape with tvecs
        campos = np.zeros((3, 1), dtype=float)
#        for i in range(tvecs.shape[0]):
        r3 = cv.Rodrigues(rvecs[0])[0]
        r4 = np.eye(4, dtype=float)
        r4[0:3,0:3] = r3.copy()
        r4[0:3,3] = tvecs[0].reshape((3, 1))[:, 0]
        r4inv = np.linalg.inv(r4)
        campos = r4inv[0:3,3].reshape((3, 1))
        edCampos.delete(0., tk.END)
        edCampos.insert(0., stringFromNp(campos, sep='\t'))
        # project points 
        prjPointsValid, jacobian = cv.projectPoints(objPoints.reshape((-1,3)), 
                                                    rvecs[0], tvecs[0], cmat, dvec)
        prjPointsValid = prjPointsValid.reshape((-1,2))
        prjPointsAll, jacobian = cv.projectPoints(points3d.reshape((-1,3)), 
                                                    rvecs[0], tvecs[0], cmat, dvec)
        prjPointsAll = prjPointsAll.reshape((-1,2))
        # calculate projection errors
        imgPointsValid = imgPoints.reshape((-1, 2))
        prjErrorsValid = (prjPointsValid - imgPointsValid)
        # display
#        prjPointsAll = newArrayByMapping(prjPointsValid, idxAllToValid)
        edCoordPrj.delete(0., tk.END)
        edCoordPrj.insert(0., stringFromNp(prjPointsAll, sep='\t'))
        prjErrorsAll = newArrayByMapping(prjErrorsValid, idxAllToValid)
        edPrjErrors.delete(0., tk.END)
        edPrjErrors.insert(0., stringFromNp(prjErrorsAll, sep='\t'))
        # all-in-one-column:rvec/tvec/cmat9/dvec14/empty/campos
        vec31 = np.zeros((31, 1), dtype=float)
        vec31[0:2, 0] = imgSize[0:2]
        vec31[2:5, 0] = rvecs[0].reshape((-1, 1))[0:3, 0]
        vec31[5:8, 0] = tvecs[0].reshape((-1, 1))[0:3, 0]
        vec31[8:17, 0] = cmat.reshape((9, 1))[0:9, 0]
        lenDvec = dvec.size
        vec31[17:17+lenDvec, 0] = dvec.reshape((lenDvec, 1))[0:lenDvec, 0]
        strOneCol = stringFromNp(vec31, sep='\t') + "\n\n"
        strOneCol += stringFromNp(campos.reshape((-1, 1)), sep='\t')
        edOneCol.delete(0., tk.END)
        edOneCol.insert(0., strOneCol)
        return
    # set button command function 
    btCalib.configure(command=btCalib_clicked)

    # #######################################################
    # Button of [Calibrate camera (chessboard)]
    # #######################################################
    btCalibCb = tk.Button(frame3, text='Calibrate camera (chessboard)', width=40, height=2)
    btCalibCb.pack()
    # define button command function 
    def btCalibCb_clicked():
        print("# *** Button [Calibrate camera (chessboard)] clicked:\\")
        # get object points and image points as Numpy array format
        # from text boxes (edCoord3d and edCoordImg)
        strPoints3d = edCoord3d.get(0., tk.END)
        strPoints2d = edCoordImg.get(0., tk.END)
        try:
            points3d = npFromString(strPoints3d).reshape((-1, 3))
        except:
            tkCalib_printMessage('Cannot parse 3D points (size: %d)' % 
                                 (int(points3d.size)))
        try:
            points2d = npFromString(strPoints2d).reshape((-1, 2))
        except:
            tkCalib_printMessage('Cannot parse 2D points (size: %d)' % 
                                 (int(points2d.size)))
            return

        # get number of chessboard files 
        try:
            cbFiles = edFile.get(0., tk.END).split()
            nCbFiles = len(cbFiles)
        except:
            tk.messagebox.showerror(title="Error", message="# Error: Cannot get calibration files.")
            return 

        # get chessboard parameters (nCornersX, nCornersY, dxCorners, dyCorners)
        try:
            strCbParams = edCbParams.get(0., tk.END)
            npCbParams = npFromString(strCbParams)
            nCornersX = int(npCbParams[0])
            nCornersY = int(npCbParams[1])
#            dxCorners = float(npCbParams[2])
#            dyCorners = float(npCbParams[3])
            nCbPhotos = points2d.shape[0] // (nCornersX * nCornersY)
        except:
            tk.messagebox.showerror(title="Error", message="# Error: Cannot get chessboard information.")
            return 
        # remove nan points, get intersection set of non-nan points
        #　get the mapping
        validsOfPoints3d, validsOfPoints2d, validCalibPoints, \
            idxAllToValid, idxValidToAll, validPoints3d, validPoints2d = \
            countCalibPoints(points3d, points2d)
        # check number of points 
        if validPoints3d.shape[0] != nCbPhotos * nCornersX * nCornersY:
            tkCalib_printMessage('Error: Number of calibration points (3D) (%d) does not match photos (%d) * corners (%dx%d)' % 
                                    (validPoints3d.shape[0], nCbPhotos, nCornersX, nCornersY))
            tkCalib_printMessage('       Calibration cancelled.')
            return
        if validPoints2d.shape[0] != nCbPhotos * nCornersX * nCornersY:
            tkCalib_printMessage('Error: Number of calibration points (image) (%d) does not match photos (%d) * corners (%dx%d)' % 
                                    (validPoints3d.shape[0], nCbPhotos, nCornersX, nCornersY))
            tkCalib_printMessage('       Calibration cancelled.')
            return
        # variables and parameters to run camera calibration 
        objPoints = validPoints3d.reshape((nCbPhotos, nCornersX * nCornersY, 3)).astype(np.float32)
        imgPoints = validPoints2d.reshape((nCbPhotos, nCornersX * nCornersY, 2)).astype(np.float32)
        imgSize = imgSizeFromBt().astype(int).reshape(-1)
        cmat = npFromString(edCmatGuess.get(0., tk.END)).reshape((3,3))
        dvec = npFromString(edDvecGuess.get(0., tk.END)).reshape((1,-1))
        flags = ck_clicked()
        rvecs = np.array([])
        tvecs = np.array([])
        print("Calibrating camera based on %d chessboard photos %dx%d corners" % (nCbPhotos, nCornersX, nCornersY))
        # run calibration 
        ret, cmat, dvec, rvecs, tvecs = cv.calibrateCamera(
            objPoints, imgPoints, imgSize, cmat, dvec, rvecs, tvecs, flags) 
        # reshape the results (rvecs and tvecs to 3d arrays)
#        rvecs = npFromTupleNp(rvecs[0]).reshape((3,1))
#        tvecs = npFromTupleNp(tvecs[0]).reshape((3,1))
        # display result to text boxes
        edRvecs.delete(0., tk.END)
        edRvecs.insert(0., stringFromNp(np.array(rvecs).reshape(-1,3), sep='\t'))
        edTvecs.delete(0., tk.END)
        edTvecs.insert(0., stringFromNp(np.array(tvecs).reshape(-1,3), sep='\t'))
        edCmat.delete(0., tk.END)
        edCmat.insert(0., stringFromNp(cmat.reshape(3,3), sep='\t'))
        edDvec.delete(0., tk.END)
        edDvec.insert(0., stringFromNp(dvec.reshape(-1,1), sep='\t'))
        # clear edCampos, edCoordPrj, edPrjErrors, edOneCol
        # edCampos: nCbPhotos * 3
        # edCoordPrj: (nCbPhotos * nCornersX * nCornersY) * 2
        # edPrjErrors: (nCbPhotos * nCornersX * nCornersY) * 2
        # edOneCol: nCbPhotos * _____
        # calculate camera position
        #   rvecs and tvecs would be a tuple of 3x1 float array(s)
#        campos = tvecs.copy() # make campos the same type and shape with tvecs
        edCampos.delete(0., tk.END)
        edCoordPrj.delete(0., tk.END)
        edPrjErrors.delete(0., tk.END)
        edOneCol.delete(0., tk.END)
        # calculate camera position, project points, and calculate projection errors
        for icam in range(nCbPhotos):
            campos = np.zeros((3, 1), dtype=float)
            r3 = cv.Rodrigues(rvecs[icam])[0]
            r4 = np.eye(4, dtype=float)
            r4[0:3,0:3] = r3.copy()
            r4[0:3,3] = tvecs[icam].reshape((3, 1))[:, 0]
            r4inv = np.linalg.inv(r4)
            campos = r4inv[0:3,3].reshape((3, 1))
            edCampos.insert(tk.END, stringFromNp(campos.flatten(), sep='\t')+'\n')
            # project points 
            prjPointsValid, jacobian = cv.projectPoints(
                objPoints.reshape(-1, (nCornersX*nCornersY), 3)[icam].reshape((-1,3)),
                rvecs[icam], tvecs[icam], cmat, dvec)
            prjPointsValid = prjPointsValid.reshape((-1,2))
            edCoordPrj.insert(tk.END, stringFromNp(prjPointsValid, sep='\t')+'\n')
            # calculate projection errors
            imgPointsValid = imgPoints.reshape(-1, (nCornersX*nCornersY), 2)[icam].reshape((-1,2))
            prjErrorsValid = (prjPointsValid - imgPointsValid)
            edPrjErrors.insert(tk.END, stringFromNp(prjErrorsValid, sep='\t')+ '\n')
            # all-in-one-column:img_W/img_H/rvec/tvec/cmat9/dvec14
            vec31 = np.zeros((31, 1), dtype=float)
            vec31[0:2, 0] = imgSize[0:2]
            vec31[2:5, 0] = rvecs[icam].reshape((-1, 1))[0:3, 0]
            vec31[5:8, 0] = tvecs[icam].reshape((-1, 1))[0:3, 0]
            vec31[8:17, 0] = cmat.reshape((9, 1))[0:9, 0]
            lenDvec = dvec.size
            vec31[17:17+lenDvec, 0] = dvec.reshape((lenDvec, 1))[0:lenDvec, 0]
            strOneCol = stringFromNp(vec31.flatten(), sep='\t')
            edOneCol.insert(tk.END, strOneCol+'\n')
        # convert back to each-camera-a-column format 
        all_in_one_txt = edOneCol.get(0., tk.END)
        #   replace '\n' with '\t'
        all_in_one_txt = all_in_one_txt.replace('\n', ' ') 
        all_in_one_txt = all_in_one_txt.replace('\t', ' ')
        all_in_one_txt = all_in_one_txt.replace(',', ' ')
        all_in_one_txt = all_in_one_txt.replace(';', ' ')
        #   get numpy array from string
        all_in_one_mat = np.fromstring(all_in_one_txt, sep=' ').reshape((nCbPhotos, -1)).T
        #   print all_in_one_mat to edOneCol with "\t" as delimiter and "\n" as line separator
        row_strings = []
        for row in all_in_one_mat:
            row_strings.append("\t".join(row.astype(str)))
        excel_compatible_string = "\n".join(row_strings)
        edOneCol.delete(0., tk.END)
        edOneCol.insert(tk.END, excel_compatible_string)
        return
    # set button command function
    btCalibCb.configure(command=btCalibCb_clicked)

    # #######################################################
    # Button of [Draw projection image]
    # #######################################################
    btDrawPoints = tk.Button(frame3, text='Draw points and grid', width=40, height=1)
    btDrawPoints.pack()
    # #######################################################
    # Button of [Draw projection image on an undistorted image]
    # #######################################################
    btDrawPointsUndist = tk.Button(frame3, text='Draw points and grid (undistorted)', 
                                   width=40, height=1)
    btDrawPointsUndist.pack()
    # #######################################################
    # Text of grid coordinates (Xs, Ys, and Zs)
    # #######################################################
    lbGridXs = tk.Label(frame3, text='xs of grid')
    lbGridXs.pack()
    edGridXs = tk.Entry(frame3, width=40)
    edGridXs.pack()
    lbGridYs = tk.Label(frame3, text='ys of grid')
    lbGridYs.pack()
    edGridYs = tk.Entry(frame3, width=40)
    edGridYs.pack()
    lbGridZs = tk.Label(frame3, text='zs of grid')
    lbGridZs.pack()
    edGridZs = tk.Entry(frame3, width=40)
    edGridZs.pack()
    lbGridImgIdx = tk.Label(frame3, text='Index of chessboard photos (1-based index)')
    lbGridImgIdx.pack()
    edGridImgIdx = tk.Entry(frame3, width=10)
    edGridImgIdx.insert(0, '1:1')
    edGridImgIdx.pack()

    #   set initial text for demonstration
    try:
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_gridXs.npy'))
        theStr = stringFromNp(theMat, sep='\t')
        edGridXs.delete(0, tk.END)
        edGridXs.insert(0, theStr)
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_gridYs.npy'))
        theStr = stringFromNp(theMat, sep='\t')
        edGridYs.delete(0, tk.END)
        edGridYs.insert(0, theStr)
        theMat = np.loadtxt(os.path.join(os.getcwd(), 'tkCalib_init_gridZs.npy'))
        theStr = stringFromNp(theMat, sep='\t')
        edGridZs.delete(0, tk.END)
        edGridZs.insert(0, theStr)
    except:
        tkCalib_printMessage("# Warning: Cannot parse tkCalib_init_gridXs/Ys/Zs.npy")
        edGridXs.delete(0, tk.END)
        edGridYs.delete(0, tk.END)
        edGridZs.delete(0, tk.END)
    # define button command function
    def btDrawPoints_clicked(undistort=False):

        # In single-photo calibration, this function draw on a single image
        # In chessboard calibration, this function draw images defined by user in 
        #  "index of chessboard photos" 
        # imgidx_start is the starting index of the image to be drawn
        # imgidx_end is the ending index of the image to be drawn
        if rbVar.get() == 0: # single-photo calibration
            imgidx_start = 1
            imgidx_end = 1
            nPhoto = 1
        else: # chessboard calibration
            imgidx_str = edGridImgIdx.get()
            # check if imgidx_str has a colon
            if ':' in imgidx_str:
                imgidx_start = int(imgidx_str.split(':')[0])
                imgidx_end = int(imgidx_str.split(':')[1])
            else:
                imgidx_start = int(imgidx_str)
                imgidx_end = int(imgidx_str)
            # number of photos is number of files in edFile
            nPhoto = len(edFile.get(0., tk.END).split())
            if imgidx_end > nPhoto:
                imgidx_end = nPhoto
            print("# imgidx_start = %d, imgidx_end = %d" % (imgidx_start, imgidx_end))
        # check if imgidx_start and imgidx_end are valid
        if imgidx_start < 1 or imgidx_end < 1:
            tkCalib_printMessage("# Error: Invalid index of chessboard photos.")
            return
        # start the loop of photos
        for imgidx in range(imgidx_start, imgidx_end+1):
            # check ratio button 
            if rbVar.get() == 0: # single-photo calibration
                print("# single-photo calibration")
            else:
                print("# chessboard calibration")
            # get background image
            try:
                fname = edFile.get(0., tk.END)
                # check if it is single-photo calibration or chessboard calibration
                if rbVar.get() == 0: # single-photo calibration
                    fname = fname.split()[0]
                else: # chessboard calibration
#                    imgidx = int(edGridImgIdx.get())
                    fname = fname.split()[imgidx-1]
                print("The background file is %s " % fname)
                bkimg = cv.imread(fname)
                if type(bkimg) != type(None) and bkimg.shape[0] > 0:
                    print("The image size is %d/%d (w/d)" % (bkimg.shape[1], bkimg.shape[0]))
                else:
                    tkCalib_printMessage("# Error: Cannot read background image %s"
                                        % bkimg)
                    raise Exception("invalid_image")
            except:
                imgSize = imgSizeFromBt()
                bkimg = np.ones((imgSize[1], imgSize[0], 3), dtype=np.uint8) * 255            
            # get intrinsic parameters
            try:
                # get cmat
                strCmat = edCmat.get(0., tk.END)
                cmat = npFromString(strCmat).reshape(3, 3)
                # get dvec
                strDvec = edDvec.get(0., tk.END)
                dvec = npFromString(strDvec).reshape(1, -1)
            except:
    #            tkCalib_printMessage("# Error: Cannot get calibrated result.")
                tk.messagebox.showerror(title="Error", message="# Error: Cannot get calibrated result..")
                return            
            # load image
            imgd = bkimg.copy()
            # get projected points
            try:
                strPrjPoints= edCoordPrj.get(0., tk.END)
                prjPointsAll = npFromString(strPrjPoints).reshape((-1, 2))
                # if it is chessboard calibration, we need to get only a part of prjPointsAll
                if rbVar.get() == 0: # single-photo calibration
                    prjPointsThisPhoto = prjPointsAll
                else: # chessboard calibration
#                    imgidx = int(edGridImgIdx.get())
                    nCornersX = int(edCbParams.get(0., tk.END).split()[0])
                    nCornersY = int(edCbParams.get(0., tk.END).split()[1])
                    prjPointsThisPhoto = prjPointsAll.reshape(-1, nCornersX*nCornersY, 2)[imgidx-1]
                if prjPointsThisPhoto.shape[0] <= 0:
                    raise Exception('invalid_proPoints')
                # draw image points
                color=[0,255,255]; 
                markerType=cv.MARKER_SQUARE
                markerSize=max(imgd.shape[0]//100, 3)
                thickness=max(imgd.shape[0]//400, 1)
                lineType=8
                imgd = drawPoints(imgd, prjPointsThisPhoto, color=color, markerType=markerType, 
                                markerSize=markerSize,
                                thickness=thickness, lineType=lineType, savefile='.')
            except:
                tkCalib_printMessage('# No projected points. Skipping plotting them.')
            # draw grid mesh (3D) on the undistorted image
            try:
                strRvec = edRvecs.get(0., tk.END)
                strTvec = edTvecs.get(0., tk.END)
                if rbVar.get() == 0: # single-photo calibration
                    rvec = npFromString(strRvec).reshape(3, -1)
                    tvec = npFromString(strTvec).reshape(3, -1)
                else:
                    rvec = npFromString(strRvec).reshape(-1, 3)[imgidx-1].reshape(3, -1)
                    tvec = npFromString(strTvec).reshape(-1, 3)[imgidx-1].reshape(3, -1)
                # get grid Xs Ys Zs
                gridXs = npFromString(edGridXs.get()).reshape((-1)).astype(np.float64)
                gridYs = npFromString(edGridYs.get()).reshape((-1)).astype(np.float64)
                gridZs = npFromString(edGridZs.get()).reshape((-1)).astype(np.float64)
                if gridXs.size > 0 and gridYs.size > 0 and gridZs.size > 0:
                    imgd = draw3dMesh(img=imgd, cmat=cmat, dvec=dvec,
                                    rvec=rvec, tvec=tvec,
                                    meshx=gridXs, meshy=gridYs, meshz=gridZs, 
                                    color=(64,255,255), thickness=1, shift=0, 
                                    savefile='.')
            except:
                print("# Skipping plotting grid.")
            # get image points
            try:
                strPointsImg= edCoordImg.get(0., tk.END)
                if rbVar.get() == 0: # single-photo calibration
                    points2d = npFromString(strPointsImg).reshape((-1, 2))
                else:                # chessboard calibration
#                    imgidx = int(edGridImgIdx.get())
                    nCornersX = int(edCbParams.get(0., tk.END).split()[0])
                    nCornersY = int(edCbParams.get(0., tk.END).split()[1])
                    points2d = npFromString(strPointsImg).reshape(-1, nCornersX*nCornersY, 2)[imgidx-1].reshape(-1, 2)
                if points2d.shape[0] <= 0:
                    raise Exception('invalid_image_points')
                # draw image points
                color=[0,255,0]; 
                markerType=cv.MARKER_CROSS
                markerSize=max(4, imgd.shape[0] // 100) # 40
                thickness=max(1, imgd.shape[0] // 400) # 10
                lineType=8
                imgd = drawPoints(imgd, points2d, color=color, markerType=markerType, 
                                markerSize=markerSize,
                                thickness=thickness, lineType=lineType, savefile='.')
            except:
                tkCalib_printMessage('# No image point. Skipping plotting image points.')
            # if undistort is true, undistort the image imgd
            if undistort:
                img_ud = cv.undistort(imgd, cmat, dvec)
                img_show = img_ud
            else:
                img_show = imgd
            # show drawn image on the screen 
    #        winW = win.winfo_width()
    #        winH = win.winfo_height()
    #        imgr = cv.resize(imgd, (winW, winH))
    #        cv.imshow("Points", imgr); cv.waitKey(0); 
            imshow2("Points", img_show, winmax=(1600, 800), interp=cv.INTER_LANCZOS4)
            try:
                cv.destroyWindow("Points")
            except:
                None
            # save the image to file
            # the file is located under a directory named tkCalib_reprojection_images
            # if this directory does not exist, create it
            fname = edFile.get(0., tk.END)
            fname_first = fname.split()[0]
            initDir = os.path.split(fname_first)[0]
            fname_this = fname.split()[imgidx-1]
            fname_this_base = os.path.split(fname_this)[1]
            try:
                os.makedirs(os.path.join(initDir, 'tkCalib_reprojection_images'))
            except:
                None
            theDir = os.path.join(initDir, 'tkCalib_reprojection_images')
            if undistort:
                # if undistort is true, save the image as "undistort_<fname_first_base>"
                theFullPath = os.path.join(theDir, 'undistort_' + fname_this_base)
            else:
                # if undistort is false, save the image as "original_<fname>"
                theFullPath = os.path.join(theDir, 'original_' + fname_this_base)
            # save image to the directory
            cv.imwrite(theFullPath, img_show)

        # # ask if user wants to save the image to file or not
        # fname = edFile.get(0., tk.END)
        # fname = fname.split()[0]
        # initDir = os.path.split(fname)[0]
        # ufileDirFile = uiputfile("Save image to file ...", initialDirectory=initDir)
        # # if xfile is selected
        # if (type(ufileDirFile) == tuple or type(ufileDirFile) == list) and \
        #     len(ufileDirFile) >= 2 and type(ufileDirFile[0]) == str and \
        #     len(ufileDirFile[1]) >= 1:
        #     # save image to selected file
        #     ufile = os.path.join(ufileDirFile[0], ufileDirFile[1])
        #     cv.imwrite(ufile, img_show)        
        return
    # set button command function
    btDrawPoints.configure(command=btDrawPoints_clicked)
    
    # define button command function
    def btDrawPointsUndist_clicked():
        btDrawPoints_clicked(undistort=True)


    # set button command function
    btDrawPointsUndist.configure(command=btDrawPointsUndist_clicked)
    # #######################################################
    # Button of [Save camera parameters (rvec/tvec/fx/fy/cx/cy/k1.../tauy)]
    # #######################################################
    btSaveCameraParameters = tk.Button(frame3, text='Save camera parameters', 
                                       width=40, height=1)
    btSaveCameraParameters.pack()
    # define button command function 
    def btSaveCameraParameters_clicked():
        # get parameters
        try:
            # image size
            strImgSize = edImgSize.get(0., tk.END)
            imgSize = npFromString(strImgSize)
            # get rvec
            strRvec = edRvecs.get(0., tk.END)
            rvec = npFromString(strRvec).reshape(3, -1)
            # get tvec
            strTvec = edTvecs.get(0., tk.END)
            tvec = npFromString(strTvec).reshape(3, -1)
            # get cmat
            strCmat = edCmat.get(0., tk.END)
            cmat = npFromString(strCmat).reshape(3, 3)
            # get dvec
            strDvec = edDvec.get(0., tk.END)
            dvec = npFromString(strDvec).reshape(1, -1)
        except:
            tkCalib_printMessage('# Error: btSaveCameraParameters_clicked: Cannot get parameters from edit texts')
        # save to file
        # ask if user wants to save the image to file or not
        fname = edFile.get(0., tk.END)
        # fname could be 1 file or multiple files. We get file [0]
        fname = fname.split()[0]
        initDir = os.path.split(fname)[0]
        ufileDirFile = uiputfile("Save camera parameters to file ...", initialDirectory=initDir)
        # if file is selected
        if (type(ufileDirFile) == tuple or type(ufileDirFile) == list) and \
            len(ufileDirFile) >= 2 and type(ufileDirFile[0]) == str and \
            len(ufileDirFile[1]) >= 1:
            # if rvec and tvec have multiple vectors (i.e., chessboard calibration)
            if rvec.shape[1] > 1 and tvec.shape[1] > 1:
                # get the Index of chessboard photos (1-based index) (edGridImgIdx)
                try:
                    imgidx = int(edGridImgIdx.get())
                except:
                    tkCalib_printMessage('# Warning: Cannot get index of chessboard photos. Set to 1 (first photo)')
                    imgidx = 1
                # get rvec and tvec of the selected photo
                rvec_save_to_file = rvec[:, imgidx-1].reshape(3, -1)
                tvec_save_to_file = tvec[:, imgidx-1].reshape(3, -1)
            else:
                # get rvec and tvec of the selected photo
                rvec_save_to_file = rvec.reshape(3, -1)
                tvec_save_to_file = tvec.reshape(3, -1)
            # get the file name
            # save image to selected file
            ufile = os.path.join(ufileDirFile[0], ufileDirFile[1])
            writeCamera(ufile, imgSize, rvec_save_to_file, tvec_save_to_file, cmat, dvec)
        #
        return
    # set button command function 
    btSaveCameraParameters.configure(command=btSaveCameraParameters_clicked)
    
    # #######################################################
    # Button of [Load calibration information from .xlsx file]
    # #######################################################
    btLoadCalibInfoFromXlsx = tk.Button(
        frame3,
        text='Load calibration information from .xlsx file',
        width=40,
        height=1
    )
    btLoadCalibInfoFromXlsx.pack()

    # define button command function
    def btLoadCalibInfoFromXlsx_clicked():
        # pop up a tk file dialog and ask user to select an xlsx file to read
        initDir = os.getcwd()
        from tkinter.filedialog import askopenfilename
        xlsxFile = askopenfilename(
            title="Load calibration information from .xlsx file",
            initialdir=initDir, defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not xlsxFile:
            print("# User cancelled loading calibration information from .xlsx file.")
            return
        # # by using openpyxl, open an existing excel file for reading
        import openpyxl
        from openpyxl import Workbook
        wb = openpyxl.load_workbook(xlsxFile)
        # read the worksheet named "calib mode"
        # skip the header line, read the 2nd line, skip the first column of 2nd line, read the rbVar
        ws = wb["calib mode"]
        calib_mode = int(ws.cell(row=2, column=2).value)
        if calib_mode == 0:
            print("# Calibration mode: single-photo calibration")
        elif calib_mode == 1:
            print("# Calibration mode: chessboard calibration")
        else:
            print("# Warning: unknown calibration mode: %d" % calib_mode)
        rbVar.set(calib_mode)    
        # read the worksheet named "Object points (3D)"
        ws = wb["Object points (3D)"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write points3d to edCoord3d
        points3d = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            points3d.append(row)
        points3d = np.array(points3d, dtype=np.float32)
        print("# Number of 3D points: %d" % points3d.shape[0])
        strPoints3d = stringFromNp(points3d, sep='\t')
        edCoord3d.delete(0., tk.END)
        edCoord3d.insert(0., strPoints3d)
        # read the worksheet named "Image points (2D)"
        ws = wb["Image points (2D)"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write points2d to edCoordImg
        points2d = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            points2d.append(row)
        points2d = np.array(points2d, dtype=np.float32)
        print("# Number of 2D points: %d" % points2d.shape[0])
        strPoints2d = stringFromNp(points2d, sep='\t')
        edCoordImg.delete(0., tk.END)
        edCoordImg.insert(0., strPoints2d)
        # read the worksheet named "Calibration image files"
        ws = wb["Calibration image files"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write file names to edFile
        file_names = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            file_names.append(row[0])
        file_names = [fname.strip() for fname in file_names if fname is not None]
        print("# Number of calibration image files: %d" % len(file_names))
        strFileNames = '\n'.join(file_names)
        edFile.delete(0., tk.END)
        edFile.insert(0., strFileNames)
        # read the worksheet named "calibration image size"
        ws = wb["Image size"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write image size to edImgSize
        imgSize = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            imgSize.append(row)
        imgSize = np.array(imgSize, dtype=np.int32).flatten()
        print("# Image size: %d x %d" % (imgSize[0], imgSize[1]))
        strImgSize = stringFromNp(imgSize, sep='\t')
        edImgSize.delete(0., tk.END)
        edImgSize.insert(0., strImgSize)
        # read the worksheet named "camera matrix"
        ws = wb["Init guess of camera matrix"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write cmat to edCmatGuess
        cmatGuess = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            cmatGuess.append(row)
        cmatGuess = np.array(cmatGuess, dtype=np.float32).reshape((3, 3))
        print("# Initial guess of camera matrix:")
        print(cmatGuess)
        strCmat = stringFromNp(cmatGuess, sep='\t')
        edCmatGuess.delete(0., tk.END)
        edCmatGuess.insert(0., strCmat)
        # read the worksheet named "Init guess of distortion"
        ws = wb["Init guess of distortion"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write dvec to edDvecGuess
        dvecGuess = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            dvecGuess.append(row)
        dvecGuess = np.array(dvecGuess, dtype=np.float32).flatten()
        print("# Initial guess of distortion vector:")
        print(dvecGuess)
        strDvec = stringFromNp(dvecGuess, sep='\t')
        edDvecGuess.delete(0., tk.END)
        edDvecGuess.insert(0., strDvec)
        # read the worksheet named "Chessboard"
        ws = wb["Chessboard"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write cbParams to edCbParams
        cbParams = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            cbParams.append(row)
        cbParams = np.array(cbParams, dtype=np.float32).flatten()
        print("# Chessboard parameters: %s" % cbParams)
        edCbParams.configure(state='normal')
        edCbParams.delete(0., tk.END)
        edCbParams.insert(0., '%d %d   %.5f   %.5f' % (cbParams[0], cbParams[1], cbParams[2], cbParams[3]))
        # read the worksheet named "Calibration flags"
        ws = wb["Calibration flags"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and update the widget by changing calibFlags
        #   calibFlags is a 1D array of uint8
        calibFlags = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            calibFlags.append(row)
        calibFlags = np.array(calibFlags, dtype=int).flatten()
        for i in range(min(len(calibFlags), len(ckValues))):
            ckValues[i].set(calibFlags[i])
        # read the worksheet named "Grid Xs"
        ws = wb["Grid Xs"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write gridXs to edGridXs
        gridXs = [] 
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            gridXs.append(row)
        gridXs = np.array(gridXs, dtype=np.float32).flatten()
        print("# Grid Xs: %s" % gridXs)
        strGridXs = stringFromNp(gridXs, sep='\t')
        edGridXs.delete(0, tk.END)
        edGridXs.insert(0, strGridXs)
        # read the worksheet named "Grid Ys"
        ws = wb["Grid Ys"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write gridYs to edGridYs
        gridYs = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            gridYs.append(row)
        gridYs = np.array(gridYs, dtype=np.float32).flatten()
        print("# Grid Ys: %s" % gridYs)
        strGridYs = stringFromNp(gridYs, sep='\t')
        edGridYs.delete(0, tk.END)
        edGridYs.insert(0, strGridYs)
        # read the worksheet named "Grid Zs"
        ws = wb["Grid Zs"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write gridZs to edGridZs
        gridZs = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            gridZs.append(row)
        gridZs = np.array(gridZs, dtype=np.float32).flatten()
        print("# Grid Zs: %s" % gridZs)
        strGridZs = stringFromNp(gridZs, sep='\t')
        edGridZs.delete(0, tk.END)
        edGridZs.insert(0, strGridZs)
        # read the worksheet named "Index of chessboard photos"
        ws = wb["Index of chessboard photos"]
        #   read the string at cell(1,1)
        strGridImgIdx = ws.cell(row=1, column=1).value
        if strGridImgIdx is None:
            strGridImgIdx = '1:1'
        #   write the string to edGridImgIdx
        edGridImgIdx.configure(state='normal')
        edGridImgIdx.delete(0, tk.END)
        edGridImgIdx.insert(0, strGridImgIdx)
        if rbVar.get() == 0: # single-photo calibration
            edGridImgIdx.configure(state='disabled')
        # read the worksheet named "Projected 2D coordinates"
        # and write data to edCoordPrj
        ws = wb["Projected 2D coordinates"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write projected points to edCoordPrj
        prjPoints = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            prjPoints.append(row)
        prjPoints = np.array(prjPoints, dtype=np.float32).reshape((-1, 2))
        print("# Number of projected points: %d" % prjPoints.shape[0])
        strPrjPoints = stringFromNp(prjPoints, sep='\t')
        edCoordPrj.delete(0., tk.END)
        edCoordPrj.insert(0., strPrjPoints)
        # read the worksheet named "Projected errors"
        #    and write data to edPrjErrors
        ws = wb["Projected errors"]
        projErrs = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            projErrs.append(row)
        projErrs = np.array(projErrs, dtype=np.float32).flatten()
        print("# Number of projected errors: %d" % projErrs.shape[0])
        strProjErrs = stringFromNp(projErrs, sep='\t')
        edPrjErrors.delete(0., tk.END)
        edPrjErrors.insert(0., strProjErrs)
        # read the worksheet named "Rvecs (calculated)"
        ws = wb["Rvecs (calculated)"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write rvecs to edRvecs
        rvecs = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            rvecs.append(row)
        rvecs = np.array(rvecs, dtype=np.float32).reshape((-1, 3))
        print("# Number of rvecs: %d" % rvecs.shape[0])
        strRvecs = stringFromNp(rvecs, sep='\t')
        edRvecs.delete(0., tk.END)
        edRvecs.insert(0., strRvecs)
        # read the worksheet named "Tvecs (calculated)"
        ws = wb["Tvecs (calculated)"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write tvecs to edTvecs
        tvecs = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            tvecs.append(row)
        tvecs = np.array(tvecs, dtype=np.float32).reshape((-1, 3))
        print("# Number of tvecs: %d" % tvecs.shape[0])
        strTvecs = stringFromNp(tvecs, sep='\t')
        edTvecs.delete(0., tk.END)
        edTvecs.insert(0., strTvecs)
        # read the worksheet named "Camera mat (calculated)"
        ws = wb["Camera matrix (calculated)"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write cmat to edCmat
        cmat = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            cmat.append(row)
        cmat = np.array(cmat, dtype=np.float32).reshape((3, 3))
        print("# Camera matrix (calculated):")
        print(cmat)
        strCmat = stringFromNp(cmat, sep='\t')
        edCmat.delete(0., tk.END)
        edCmat.insert(0., strCmat)
        # read the worksheet named "Distortion coeff. (calculated)"
        ws = wb["Distortion coeff. (calculated)"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write dvec to edDvec
        dvec = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            dvec.append(row)
        dvec = np.array(dvec, dtype=np.float32).flatten()
        print("# Number of distortion coefficients: %d" % dvec.shape[0])
        strDvec = stringFromNp(dvec, sep='\t')
        edDvec.delete(0., tk.END)
        edDvec.insert(0., strDvec)
        # read the worksheet named "Camera positions (calculated)"
        ws = wb["Camera positions (calculated)"]
        #   skip the header line, read the rest of the lines but skip the first column of each line
        #   and write camera positions to edCampos
        camPos = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            camPos.append(row)
        camPos = np.array(camPos, dtype=np.float32).reshape((-1, 3))
        print("# Number of camera positions: %d" % camPos.shape[0])
        strCamPos = stringFromNp(camPos, sep='\t')
        edCampos.delete(0., tk.END)
        edCampos.insert(0., strCamPos)
        # read the worksheet named "All-in-one"
        #   read a 2D array with arbitrary size from cell(1,1)
        #     skip the header line, read the rest of the lines but skip the first column of each line
        ws = wb["parameters-all-in-one"]
        all_in_one_data = []
        for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
            all_in_one_data.append(row)
        n_rows = len(all_in_one_data)
        all_in_one_mat = np.array(all_in_one_data, dtype=np.float32)
        all_in_one_mat = all_in_one_mat.reshape((n_rows, -1))
        print("# All-in-one data: %s" % all_in_one_data)
        #   write all-in-one data to edOneCol
        strAllInOne = stringFromNp(all_in_one_mat, sep='\t')
        edOneCol.delete(0., tk.END)
        edOneCol.insert(0., strAllInOne)
        # close the workbook
        wb.close()
        return 
    

    # set button command function
    btLoadCalibInfoFromXlsx.configure(command=btLoadCalibInfoFromXlsx_clicked)

    # #######################################################
    # Button of [Save calibration information to .xlsx file]
    # #######################################################
    btSaveCalibInfoToXlsx = tk.Button(
        frame3,
        text='Save calibration information to .xlsx file',
        width=40,
        height=1
    )
    btSaveCalibInfoToXlsx.pack()

    # define button command function
    def btSaveCalibInfoToXlsx_clicked():
        import re
        delimiter_pattern = r'[ ,;:\s|\t]+' 
        # pop up a tk file dialog and ask user to select an xlsx file to save
        initDir = os.getcwd()
        from tkinter.filedialog import asksaveasfilename
        xlsxFile = asksaveasfilename(
            title="Save calibration information to .xlsx file",
            initialdir=initDir,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not xlsxFile:
            print("# User cancelled saving calibration information to .xlsx file.")
            return
        # by using openpyxl, open a new excel file for writing
        import openpyxl
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # calibration mode (0:single-photo, 1:chessboard)
        calib_mode = int(rbVar.get())
        if calib_mode == 0:
            print("# Calibration mode: single-photo calibration")
        elif calib_mode == 1:
            print("# Calibration mode: chessboard calibration")
        else:
            print("# Warning: unknown calibration mode: %d" % calib_mode) 
        #   create a worksheet named "calib mode". Header row is "key" and "value", 
        #   2nd row is "Calibration mode (0:single-photo, 1:chessboard)", and integer of calib_mode
        ws.title = "calib mode"
        ws.append(["key", "value"])
        ws.append(["Calibration mode (0:single-photo, 1:chessboard)", calib_mode])
        # text in 3D coordinates (edCoord3d)
        strPoints3d = edCoord3d.get(0., tk.END)
        strPoints3d = re.sub(delimiter_pattern, ' ', strPoints3d)
        points3d = np.fromstring(strPoints3d, sep=' ', dtype=float).reshape((-1, 3))
        print("# Number of 3D points: %d" % points3d.shape[0])
        #   create a worksheet named "Object (3D) points in world coord". Header row is "id", "xw", "yw", "zw",
        #   and 3D coordinates of points
        ws = wb.create_sheet(title="Object points (3D)")
        ws.append(["object point id (1-base)", "x", "y", "z"]) 
        for i in range(points3d.shape[0]):
            ws.append([i+1, points3d[i, 0], points3d[i, 1], points3d[i, 2]])
        # text in 2D coordinates (edCoordImg)
        strPointsImg = edCoordImg.get(0., tk.END)
        strPointsImg = re.sub(delimiter_pattern, ' ', strPointsImg)
        pointsImg = np.fromstring(strPointsImg, sep=' ', dtype=float).reshape((-1, 2))
        print("# Number of (2D) image points: %d" % pointsImg.shape[0])
        #   create a worksheet named "Image (2D) points in image coord". Header row is "id (1-base)", "xi", "yi",
        #   and 2D coordinates of points
        ws = wb.create_sheet(title="Image points (2D)")
        ws.append(["image point id (1-base)", "xi", "yi"])
        for i in range(pointsImg.shape[0]):
            ws.append([i+1, pointsImg[i, 0], pointsImg[i, 1]])
        # text in select calibration image file (edFile)
        strCalibImg = edFile.get(0., tk.END)
        strCalibImg = re.sub(r'[ ,;\s|\t]+' , ' ', strCalibImg)
        strCalibImg = strCalibImg.split()
        print("# Number of calibration image files: %d" % len(strCalibImg))
        #   create a worksheet named "Calibration image files". Header row is "calibration photo file id (1-base)", "file name",
        #   and file names of calibration images
        ws = wb.create_sheet(title="Calibration image files")
        ws.append(["calibration image file id (1-base)", "file name"])
        for i in range(len(strCalibImg)):
            fname = strCalibImg[i]
            # check if fname is a valid file
            ws.append([i+1, fname])
        # text in image size (edImgSize)
        strImgSize = edImgSize.get(0., tk.END)
        strImgSize = re.sub(delimiter_pattern, ' ', strImgSize)
        imgSize = np.fromstring(strImgSize, sep=' ', dtype=int).reshape((-1, 2))
        img_width = int(imgSize[0, 0])
        img_height = int(imgSize[0, 1])
        print("# Image width/height: %d/%d" % (img_width, img_height))
        #   create a worksheet named "Image size". Header row is "key", "value"
        ws = wb.create_sheet(title="Image size")
        ws.append(["key", "value"])
        ws.append(["Image width", img_width])
        ws.append(["Image height", img_height])
        # text in camera mat (init guess) (edCmatGuess)
        strCmatGuess = edCmatGuess.get(0., tk.END)
        strCmatGuess = re.sub(delimiter_pattern, ' ', strCmatGuess)
        cmatGuess = np.fromstring(strCmatGuess, sep=' ', dtype=float).reshape((3, 3))
        print("# Initial guess of camera matrix: \n%s" % cmatGuess)
        #    create a worksheet named "Initial guess of camera matrix". Header row is "key", "value"
        ws = wb.create_sheet(title="Init guess of camera matrix")
        ws.append(["key", "value"])
        ws.append(["cmat_11 (focal length x)", cmatGuess[0, 0]])
        ws.append(["cmat_12", cmatGuess[0, 1]])
        ws.append(["cmat_13 (principal point x)", cmatGuess[0, 2]])
        ws.append(["cmat_21", cmatGuess[1, 0]])
        ws.append(["cmat_22 (focal length y)", cmatGuess[1, 1]])
        ws.append(["cmat_23 (principal point y)", cmatGuess[1, 2]])
        ws.append(["cmat_31", cmatGuess[2, 0]])
        ws.append(["cmat_32", cmatGuess[2, 1]])
        ws.append(["cmat_33", cmatGuess[2, 2]])
        # text in distortion vector (init guess) (edDvecGuess)
        strDvecGuess = edDvecGuess.get(0., tk.END)
        strDvecGuess = re.sub(delimiter_pattern, ' ', strDvecGuess)
        dvecGuess = np.fromstring(strDvecGuess, sep=' ', dtype=float).reshape((1, -1))
        print("# Initial guess of distortion vector: \n%s" % dvecGuess)
        #   create a worksheet named "Initial guess of distortion vector". Header row is "key", "value"
        ws = wb.create_sheet(title="Init guess of distortion")
        ws.append(["key", "value"])
        ws.append(["k1", dvecGuess[0, 0]])
        ws.append(["k2", dvecGuess[0, 1]])
        ws.append(["p1", dvecGuess[0, 2]])
        ws.append(["p2", dvecGuess[0, 3]])
        if dvecGuess.shape[1] > 4:
            ws.append(["k3", dvecGuess[0, 4]])
        if dvecGuess.shape[1] > 5:
            ws.append(["k4", dvecGuess[0, 5]])
        if dvecGuess.shape[1] > 6:
            ws.append(["k5", dvecGuess[0, 6]])
        if dvecGuess.shape[1] > 7:
            ws.append(["k6", dvecGuess[0, 7]])
        if dvecGuess.shape[1] > 8:
            ws.append(["s1", dvecGuess[0, 8]])
        if dvecGuess.shape[1] > 9:
            ws.append(["s2", dvecGuess[0, 9]])
        if dvecGuess.shape[1] > 10:
            ws.append(["s3", dvecGuess[0, 10]])
        if dvecGuess.shape[1] > 11:
            ws.append(["s4", dvecGuess[0, 11]])
        if dvecGuess.shape[1] > 12:
            ws.append(["tau_x", dvecGuess[0, 12]])
        if dvecGuess.shape[1] > 13:
            ws.append(["tau_y", dvecGuess[0, 13]])
        # text in find chessboard corners parameters (edCbParams)
        strCbParams = edCbParams.get(0., tk.END)
        strCbParams = re.sub(delimiter_pattern, ' ', strCbParams)
        cbParams = np.fromstring(strCbParams, sep=' ', dtype=float).reshape((-1, 4))
        #   create a worksheet named "Find chessboard corners parameters". Header row is "key", "value"
        ws = wb.create_sheet(title="Chessboard")
        ws.append(["key", "value"])
        ws.append(["nCornersX", int(cbParams[0, 0])])
        ws.append(["nCornersY", int(cbParams[0, 1])])
        ws.append(["squareSizeX (the same unit of 3D object points, e.g., mm)", cbParams[0, 2]])
        ws.append(["squareSizeY (the same unit of 3D object points, e.g., mm)", cbParams[0, 3]])
        # values in check buttons (ckValues) (flags for calibration) 
        current_flags = [ckValues[kk].get() for kk in range(len(ckValues))]
        #   create a worksheet named "Calibration flags". Header row is "key", "value"
        ws = wb.create_sheet(title="Calibration flags")
        ws.append(["key", "value"])
        for kk in range(min(len(ckValues), len(strFlags))):
            ws.append([strFlags[kk], current_flags[kk]])
        # xs in grid (edGridXs)
        strGridXs = edGridXs.get()
        strGridXs = re.sub(delimiter_pattern, ' ', strGridXs)
        gridXs = np.fromstring(strGridXs, sep=' ', dtype=float).reshape((-1))
        print("# Number of grid Xs: %d" % gridXs.shape[0])
        #   create a worksheet named "Grid Xs". Header row is "id", "x"
        ws = wb.create_sheet(title="Grid Xs")
        ws.append(["grid X id (1-base)", "x"])
        for i in range(gridXs.shape[0]):
            ws.append([i+1, gridXs[i]])
        # ys in grid (edGridYs)
        strGridYs = edGridYs.get()
        strGridYs = re.sub(delimiter_pattern, ' ', strGridYs)
        gridYs = np.fromstring(strGridYs, sep=' ', dtype=float).reshape((-1))
        print("# Number of grid Ys: %d" % gridYs.shape[0])
        #   create a worksheet named "Grid Ys". Header row is "id", "y"
        ws = wb.create_sheet(title="Grid Ys")
        ws.append(["grid Y id (1-base)", "y"])
        for i in range(gridYs.shape[0]):
            ws.append([i+1, gridYs[i]])
        # zs in grid (edGridZs)
        strGridZs = edGridZs.get()
        strGridZs = re.sub(delimiter_pattern, ' ', strGridZs)
        gridZs = np.fromstring(strGridZs, sep=' ', dtype=float).reshape((-1))
        print("# Number of grid Zs: %d" % gridZs.shape[0])
        #   create a worksheet named "Grid Zs". Header row is "id", "z"
        ws = wb.create_sheet(title="Grid Zs")
        ws.append(["grid Z id (1-base)", "z"])
        for i in range(gridZs.shape[0]):
            ws.append([i+1, gridZs[i]])
        # index of chessboard photos (edGridImgIdx) to show with grid
        strGridImgIdx = edGridImgIdx.get()
        #   create a worksheet named "Index of chessboard photos to show with grid". Header row is "key", "value"
        ws = wb.create_sheet(title="Index of chessboard photos")
        #   append the string strGridImgIdx
        ws.append([strGridImgIdx])        
        # create a worksheet named "Projected 2D coordinates". Header row is "id", "x", "y"
        ws = wb.create_sheet(title="Projected 2D coordinates")
        ws.append(["id", "xi", "yi"])
        strPrjPoints = edCoordPrj.get(0., tk.END)
        strPrjPoints = re.sub(delimiter_pattern, ' ', strPrjPoints)
        prjPoints = np.fromstring(strPrjPoints, sep=' ', dtype=float).reshape((-1, 2))
        for i in range(prjPoints.shape[0]):
            ws.append([i+1, prjPoints[i, 0], prjPoints[i, 1]])
        print("# Number of projected 2D coordinates: %d" % prjPoints.shape[0])
        # create a worksheet named "Projected errors". Header row is "id", "error_x", "error_y"
        ws = wb.create_sheet(title="Projected errors")
        ws.append(["id", "error_x", "error_y"])
        strPrjErrors = edPrjErrors.get(0., tk.END)
        strPrjErrors = re.sub(delimiter_pattern, ' ', strPrjErrors)
        prjErrors = np.fromstring(strPrjErrors, sep=' ', dtype=float).reshape((-1, 2))
        for i in range(prjErrors.shape[0]):
            ws.append([i+1, prjErrors[i, 0], prjErrors[i, 1]])
        # create a worksheet named "Rvecs (calculated)". Header row is "id", "rvec_x", "rvec_y", "rvec_z"
        ws = wb.create_sheet(title="Rvecs (calculated)")
        ws.append(["id", "rvec_x", "rvec_y", "rvec_z"])
        strRvecs = edRvecs.get(0., tk.END)
        strRvecs = re.sub(delimiter_pattern, ' ', strRvecs)
        rvecs = np.fromstring(strRvecs, sep=' ', dtype=float).reshape((-1, 3))
        print("# Number of Rvecs (calculated): %d" % rvecs.shape[0])
        for i in range(rvecs.shape[0]):
            ws.append([i+1, rvecs[i, 0], rvecs[i, 1], rvecs[i, 2]])
        # create a worksheet named "Tvecs (calculated)". Header row is "id", "tvec_x", "tvec_y", "tvec_z"
        ws = wb.create_sheet(title="Tvecs (calculated)")
        ws.append(["id", "tvec_x", "tvec_y", "tvec_z"])
        strTvecs = edTvecs.get(0., tk.END)
        strTvecs = re.sub(delimiter_pattern, ' ', strTvecs)
        tvecs = np.fromstring(strTvecs, sep=' ', dtype=float).reshape((-1, 3))
        print("# Number of Tvecs (calculated): %d" % tvecs.shape[0])
        for i in range(tvecs.shape[0]):
            ws.append([i+1, tvecs[i, 0], tvecs[i, 1], tvecs[i, 2]])
        # create a worksheet named "Camera mat (calculated)". Header row is "key", "value"
        ws = wb.create_sheet(title="Camera matrix (calculated)")
        ws.append(["key", "value"])
        strCmat = edCmat.get(0., tk.END)
        strCmat = re.sub(delimiter_pattern, ' ', strCmat)
        cmat = np.fromstring(strCmat, sep=' ', dtype=float).reshape((3, 3))
        print("# Camera matrix (calculated): \n%s" % cmat)
        # write camera matrix to the worksheet
        ws.append(["cmat_11 (focal length x)", cmat[0, 0]])
        ws.append(["cmat_12", cmat[0, 1]])
        ws.append(["cmat_13 (principal point x)", cmat[0, 2]])
        ws.append(["cmat_21", cmat[1, 0]])
        ws.append(["cmat_22 (focal length y)", cmat[1, 1]])
        ws.append(["cmat_23 (principal point y)", cmat[1, 2]])
        ws.append(["cmat_31", cmat[2, 0]])
        ws.append(["cmat_32", cmat[2, 1]])
        ws.append(["cmat_33 (principal point z)", cmat[2, 2]])
        # create a worksheet named "Distortion coeff. (calculated)". Header row is "key", "value"
        ws = wb.create_sheet(title="Distortion coeff. (calculated)")
        ws.append(["key", "value"])
        strDvec = edDvec.get(0., tk.END)
        strDvec = re.sub(delimiter_pattern, ' ', strDvec)
        dvec = np.fromstring(strDvec, sep=' ', dtype=float).reshape((1, -1))
        print("# Distortion vector (calculated): \n%s" % dvec)
        # write distortion vector to the worksheet
        ws.append(["k1", dvec[0, 0]])
        ws.append(["k2", dvec[0, 1]])
        ws.append(["p1", dvec[0, 2]])
        ws.append(["p2", dvec[0, 3]])
        if dvec.shape[1] > 4:
            ws.append(["k3", dvec[0, 4]])
        if dvec.shape[1] > 5:
            ws.append(["k4", dvec[0, 5]])
        if dvec.shape[1] > 6:
            ws.append(["k5", dvec[0, 6]])
        if dvec.shape[1] > 7:
            ws.append(["k6", dvec[0, 7]])
        if dvec.shape[1] > 8:
            ws.append(["s1", dvec[0, 8]])
        if dvec.shape[1] > 9:
            ws.append(["s2", dvec[0, 9]])
        if dvec.shape[1] > 10:
            ws.append(["s3", dvec[0, 10]])
        if dvec.shape[1] > 11:
            ws.append(["s4", dvec[0, 11]])
        if dvec.shape[1] > 12:
            ws.append(["tau_x", dvec[0, 12]])
        if dvec.shape[1] > 13:
            ws.append(["tau_y", dvec[0, 13]])
        # create a worksheet named "Camera positions (calculated)". Header row is "id", "x", "y", "z"
        ws = wb.create_sheet(title="Camera positions (calculated)")
        ws.append(["id", "x", "y", "z"])
        strCampos = edCampos.get(0., tk.END)
        strCampos = re.sub(delimiter_pattern, ' ', strCampos)
        campos = np.fromstring(strCampos, sep=' ', dtype=float).reshape((-1, 3))
        print("# Number of camera positions (calculated): %d" % campos.shape[0])
        for i in range(campos.shape[0]):
            ws.append([i+1, campos[i, 0], campos[i, 1], campos[i, 2]])
        # create a worksheet named "All-in-one". 
        ws = wb.create_sheet(title="parameters-all-in-one")
        strAllInOne = edOneCol.get(0., tk.END)
        all_in_one_mat = np.fromstring(strAllInOne, sep=' ', dtype=float).reshape(-1, tvecs.shape[0])
        #  write the header row: "parameter", "camera_1", "camera_2", ..., "camera_%d" % tvecs.shape[0]
        #  first column is "parameter", "image_width", "image_height", "rvec_x", "rvec_y", "rvec_z",
        #  "tvec_x", "tvec_y", "tvec_z", "cmat_11", "cmat_12", "cmat_13", "cmat_21", "cmat_22",
        #  "cmat_23", "cmat_31", "cmat_32", "cmat_33", "k1", "k2", "p1", "p2", "k3", "k4",
        #  "k5", "k6", "s1", "s2", "s3", "s4", "tau_x", "tau_y"
        parm_strings = ["image_width", "image_height", "rvec_x", "rvec_y", "rvec_z", "tvec_x", "tvec_y", "tvec_z",
                        "cmat_11", "cmat_12", "cmat_13", "cmat_21", "cmat_22", "cmat_23",
                        "cmat_31", "cmat_32", "cmat_33", "k1", "k2", "p1", "p2", "k3", "k4",
                        "k5", "k6", "s1", "s2", "s3", "s4", "tau_x", "tau_y"]
        header_row = ["parameter"]
        for i in range(tvecs.shape[0]):
            header_row.append("camera_%d" % (i + 1))
        ws.append(header_row)
        # append the 2d numpy matrix all_in_one_mat to the worksheet
        for irow in range(all_in_one_mat.shape[0]):
            # append parm_strings[irow] to the first column
            ws.append([parm_strings[irow]] + all_in_one_mat[irow, :].tolist())
#            ws.append(all_in_one_mat[irow, :].tolist())
        # close workbook and save the file
        # if xlsxFile is not empty, save the workbook
        wb.save(xlsxFile)
        wb.close()
        print("# Calibration information saved to %s" % xlsxFile)
        return
    # set button command function
    btSaveCalibInfoToXlsx.configure(command=btSaveCalibInfoToXlsx_clicked)

    # #######################################################
    # Frame 4
    # #######################################################
    # Button and edit text of [Projected 2D coordinates]
    # #######################################################
    btCoordPrj = tk.Button(frame4, text='Projected coordinates')
    btCoordPrj.pack()
    btCoordPrj.config(state='disabled')
    #   edit text with text height of 5
    edCoordPrj = tk.Text(frame4, width=40, height=3, undo=True, 
                         autoseparators=True, maxundo=-1)
    edCoordPrj.pack()
    #   set initial text for demonstration
    edCoordPrj.delete(0., tk.END)
    # #######################################################
    # Button and edit text of [Projected errors]
    # #######################################################
    btPrjErrors = tk.Button(frame4, text='Projected errors')
    btPrjErrors.pack()
    #   edit text with text height of 5
    edPrjErrors = tk.Text(frame4, width=40, height=3, undo=True, 
                          autoseparators=True, maxundo=-1)
    edPrjErrors.pack()
    #   set initial text for demonstration
    edPrjErrors.delete(0., tk.END)    
    # #######################################################
    # Button of [Rvecs (calculated)]
    # #######################################################
    btRvecs = tk.Button(frame4, text='Rvecs.')
    btRvecs.pack()
    edRvecs = tk.Text(frame4, width=40, height=2, undo=True, 
                      autoseparators=True, maxundo=-1)
    edRvecs.pack()
    # #######################################################
    # Button of [Tvecs (calculated)]
    # #######################################################
    # Calculated rvecs
    btTvecs = tk.Button(frame4, text='Tvecs.')
    btTvecs.pack()
    edTvecs = tk.Text(frame4, width=40, height=2, undo=True, 
                      autoseparators=True, maxundo=-1)
    edTvecs.pack()
    # #######################################################
    # Button of [Camera mat (calculated)]
    # #######################################################
    # Calculated camera matrix 
    btCmat = tk.Button(frame4, text='Camera mat')
    btCmat.pack()
    edCmat = tk.Text(frame4, width=40, height=2, undo=True, 
                     autoseparators=True, maxundo=-1)
    edCmat.pack()
    # #######################################################
    # Button of [Distortion coeff. (calculated)]
    # #######################################################
    # Calculated distortion vector 
    btDvec = tk.Button(frame4, text='Distortion coeff.')
    btDvec.pack()
    edDvec = tk.Text(frame4, width=40, height=2, undo=True, 
                     autoseparators=True, maxundo=-1)
    edDvec.pack()
    # #######################################################
    # Button of [Camera positions (calculated)]
    # #######################################################
    btCampos = tk.Button(frame4, text='Camera position(s)')
    btCampos.pack()
    edCampos = tk.Text(frame4, width=40, height=2, undo=True, 
                       autoseparators=True, maxundo=-1)
    edCampos.pack()
    # #######################################################
    # Button of [Parameters all-in-one-column]
    # #######################################################
    btOneCol = tk.Button(frame4, text='All-in-one-column')
    btOneCol.pack()
    edOneCol = tk.Text(frame4, width=40, height=2, undo=True, 
                       autoseparators=True, maxundo=-1)
    edOneCol.pack()
    # #######################################################
    # function of close()
    # #######################################################
    def winClose():
        # nonlocal
        nonlocal imgSize, rvec, tvec, cmat, dvec
        # save current edCoord3d
        try:
            strPoints3d = edCoord3d.get(0., tk.END)
            matPoints3d = npFromString(strPoints3d).reshape((-1,3))
            matPoints3d = matPoints3d.astype(np.float32)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_coord3d.npy'), matPoints3d)
        except:
            tkCalib_printMessage("# Error: Cannot parse 3D coord.")
        # save current edCoordImg
        try:
            strPointsImg= edCoordImg.get(0., tk.END)
            matPoints2d = npFromString(strPointsImg).reshape((-1,2))
            matPoints2d = matPoints2d.astype(np.float32)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_coordImg.npy'), matPoints2d)
        except:
            tkCalib_printMessage("# Error: Cannot parse image (2D) coord.")
        # save current calibration image file
        try:
            strCalibImg = edFile.get(0., tk.END)
            tfile = open(os.path.join(os.getcwd(), 'tkCalib_init_imgFile.txt'), "w")
            n = tfile.write(strCalibImg)
            tfile.close()
        except:
            tkCalib_printMessage("# Error: Cannot get calibration image file path")
        # save current image size
        try:
            theStr = edImgSize.get(0., tk.END)
            imgSize = npFromString(theStr).reshape((1,-1))
            imgSize = imgSize.astype(int)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_imgSize.npy'), imgSize, fmt='%d')
        except:
            tkCalib_printMessage("# Error: Cannot parse image size")
        # save current cmatGuess
        try:
            theStr = edCmatGuess.get(0., tk.END)
            theMat = npFromString(theStr).reshape((3,3))
            theMat = theMat.astype(np.float64)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_cmatGuess.npy'), theMat)
        except:
            tkCalib_printMessage("# Error: Cannot parse camera mat (guessed).")
        # save current dvecGuess
        try:
            theStr = edDvecGuess.get(0., tk.END)
            theMat = npFromString(theStr).reshape((-1))
            theMat = theMat.astype(np.float64)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_dvecGuess.npy'), theMat)
        except:
            tkCalib_printMessage("# Error: Cannot parse distortion vec (guessed).")
        # save current chessboard parameters
        try:
            theStr = edCbParams.get(0., tk.END)
            theMat = npFromString(theStr).reshape((-1))
            theMat = theMat.astype(np.float64)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_cboardParams.npy'), theMat)
        except:
            tkCalib_printMessage("# Error: Cannot parse calibration board parameters (nCornersX nCornersY dxCorners dyCorners.")
        # save current flags
        try:
            theStr = edFlags.get()
            theStr = theStr[theStr.find(':')+1:]
            theMat = npFromString(theStr).reshape((-1))
            theMat = theMat.astype(np.int32)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_calibFlags.npy'), theMat, fmt='%i')
        except:
            tkCalib_printMessage("# Error: Cannot parse calibration flags.")
        # save current gridXs/Ys/Zs
        try:
            theStr = edGridXs.get()
            theMat = npFromString(theStr).reshape((-1)).astype(np.float64)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_gridXs.npy'), theMat)
            theStr = edGridYs.get()
            theMat = npFromString(theStr).reshape((-1)).astype(np.float64)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_gridYs.npy'), theMat)
            theStr = edGridZs.get()
            theMat = npFromString(theStr).reshape((-1)).astype(np.float64)
            np.savetxt(os.path.join(os.getcwd(), 'tkCalib_init_gridZs.npy'), theMat)
        except:
            tkCalib_printMessage("# Error: Cannot parse grid coordinates (Xs/Ys/Zs).")
                
        print("tkCalib window closed.")
        imgSize = npFromString(edImgSize.get(0., tk.END)).astype(int).flatten()
        try:
            rvec = npFromString(edRvecs.get(0., tk.END)).reshape(3, -1)
        except:
            rvec = np.array([], dtype=float)
        try:
            tvec = npFromString(edTvecs.get(0., tk.END)).reshape(3, -1)
        except:
            tvec = np.array([], dtype=float)
        try:
            cmat = npFromString(edCmat.get(0., tk.END)).reshape((3, 3))
        except:
            cmat = np.array([], dtype=float)
        try:
            dvec = npFromString(edDvec.get(0., tk.END)).reshape((1,-1))
        except:
            dvec = np.array([], dtype=float)
        win.destroy()
        return     
    
    rb_clicked()
    win.lift()
    win.attributes("-topmost", True)
    win.after_idle(win.attributes,'-topmost',False)
    win.protocol("WM_DELETE_WINDOW", winClose)

    win.mainloop()
#    win.destroy()


if __name__ == "__main__":
#    camParams = tkCalib()
    tkCalib()